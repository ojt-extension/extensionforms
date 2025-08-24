# views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.core.paginator import Paginator
from django.db import transaction, IntegrityError
from django.urls import reverse
from django.views.generic import FormView
from django.core.exceptions import ValidationError
from django.utils.html import escape
from django.utils.safestring import mark_safe
from django.utils._os import safe_join
from .models import (
    Partnership, InterFep, ExterFep, AdvServices, LeadUnit, PartnerAgencies,
    ProjectNoInText, CmmSrvy, IfpoProg, Proj, AwardRcvd, IfpoProgExter,
    ProjExter, AwardsRcvdExter, AdvSClients, AsTimeliness, EstExsFund,
    AsOverallRating, PartnershipFile, PartnershipImage, InterFepFile,
    InterFepImage, ExterFepFile, ExterFepImage, AdvServicesFile, AdvServicesImage,
    AutoPopulationManager
)
from .forms import PartnershipForm, InterFepForm, ExterFepForm, AdvServicesForm
from json import loads as json_loads
import csv
from os.path import basename as os_basename, exists as os_exists
from typing import Any, Dict, List, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ============================================================================
# UTILITY FUNCTIONS (Moved from utils.py)
# ============================================================================

def handle_sub_field_creation(model_class, data_dict, field_name):
    """Helper function to handle sub-field data creation with error handling"""
    if any(data_dict.values()):
        try:
            obj, created = model_class.objects.get_or_create(**data_dict)
            return obj
        except (IntegrityError, ValidationError) as e:
            return None
    return None

FILE_SIZE_LIMIT = 10 * 1024 * 1024  # 10MB
MAX_FILES_PER_UPLOAD = 10

def get_field_value(obj, field_path):
    """
    Get field value from object, handling nested relationships and missing data
    """
    if not obj:
        return ''
    
    try:
        # Handle nested field paths (e.g., 'leadunit.department')
        if '.' in field_path:
            parts = field_path.split('.')
            value = obj
            for part in parts:
                if hasattr(value, part):
                    value = getattr(value, part)
                    if value is None:
                        return ''
                else:
                    return ''
            return str(value) if value is not None else ''
        else:
            # Handle direct field access
            if hasattr(obj, field_path):
                value = getattr(obj, field_path)
                if hasattr(value, 'all'):  # Handle many-to-many fields
                    return ', '.join([str(item) for item in value.all()])
                return str(value) if value is not None else ''
            else:
                return ''
    except (AttributeError, TypeError):
        return ''

def export_forms_to_excel(request):
    """
    Exports current user's form entries to Excel with filled and blank forms
    """
    if not HAS_OPENPYXL:
        return export_forms_to_csv(request)
    
    workbook = Workbook()
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=current_forms_export.xlsx'

    # Remove default sheet
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])

    # Export Partnerships
    partnerships = Partnership.objects.select_related('leadunit', 'partagency').order_by('partship_id')
    sheet = workbook.create_sheet(title='Partnerships')
    
    headers = ['Partnership ID', 'MOA Number', 'Code', 'Category', 'Date Notarized',
               'Lead Unit Department', 'Partner Agency Name', 'Nature of Partnership',
               'SDG Number', 'Thematic Area', 'Remarks']
    
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    for row, p in enumerate(partnerships, 2):
        sheet.cell(row=row, column=1, value=p.partship_id)
        sheet.cell(row=row, column=2, value=p.moano or '')
        sheet.cell(row=row, column=3, value=p.code or '')
        sheet.cell(row=row, column=4, value=p.categ or '')
        sheet.cell(row=row, column=5, value=str(p.datenotar) if p.datenotar else '')
        sheet.cell(row=row, column=6, value=p.leadunit.department if p.leadunit else '')
        sheet.cell(row=row, column=7, value=p.partagency.nameagen if p.partagency else '')
        sheet.cell(row=row, column=8, value=p.natpart or '')
        sheet.cell(row=row, column=9, value=p.sdgnum or '')
        sheet.cell(row=row, column=10, value=p.themarea or '')
        sheet.cell(row=row, column=11, value=p.remarks or '')

    # Export Internal Projects with blank forms
    internal_projects = InterFep.objects.select_related('lunitid', 'collabagenid', 'partship__partagency').all()
    sheet2 = workbook.create_sheet(title='Internal Projects')
    
    int_headers = ['Internal Project ID', 'Partnership ID', 'Partnership Agency', 'Code',
                   'Lead Unit Department', 'Collaborating Agency', 'Beneficiaries',
                   'SDG Number', 'Thematic Area', 'Remarks']
    
    for col, header in enumerate(int_headers, 1):
        cell = sheet2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    partnerships_with_internal = set()
    row = 2
    
    # Add filled forms
    for proj in internal_projects:
        sheet2.cell(row=row, column=1, value=proj.interfepid)
        sheet2.cell(row=row, column=2, value=proj.partship.partship_id if proj.partship else '')
        sheet2.cell(row=row, column=3, value=proj.partship.partagency.nameagen if proj.partship and proj.partship.partagency else '')
        sheet2.cell(row=row, column=4, value=proj.code or '')
        sheet2.cell(row=row, column=5, value=proj.lunitid.department if proj.lunitid else '')
        sheet2.cell(row=row, column=6, value=proj.collabagenid.nameagen if proj.collabagenid else '')
        sheet2.cell(row=row, column=7, value='')
        sheet2.cell(row=row, column=8, value=proj.sdgnum or '')
        sheet2.cell(row=row, column=9, value=proj.themarea or '')
        sheet2.cell(row=row, column=10, value=proj.remarks or '')
        
        if proj.partship:
            partnerships_with_internal.add(proj.partship.partship_id)
        row += 1
    
    # Add blank forms
    internal_partnerships = Partnership.objects.filter(natpart__in=['Internal', 'Both'])
    for p in internal_partnerships:
        if p.partship_id not in partnerships_with_internal:
            sheet2.cell(row=row, column=1, value='')
            sheet2.cell(row=row, column=2, value=p.partship_id)
            sheet2.cell(row=row, column=3, value=p.partagency.nameagen if p.partagency else '')
            sheet2.cell(row=row, column=4, value='')
            sheet2.cell(row=row, column=5, value=p.leadunit.department if p.leadunit else '')
            sheet2.cell(row=row, column=6, value=p.partagency.nameagen if p.partagency else '')
            sheet2.cell(row=row, column=7, value='')
            sheet2.cell(row=row, column=8, value=p.sdgnum or '')
            sheet2.cell(row=row, column=9, value=p.themarea or '')
            sheet2.cell(row=row, column=10, value='')
            row += 1

    workbook.save(response)
    return response

def export_forms_to_csv(request):
    """
    Fallback CSV export function when openpyxl is not available
    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="current_forms_export.csv"'
    
    writer = csv.writer(response)
    
    # Export partnerships
    partnerships = Partnership.objects.select_related('leadunit', 'partagency').order_by('partship_id')
    
    # Write partnership headers
    partnership_headers = [
        'Partnership ID', 'MOA Number', 'Code', 'Category', 'Date Notarized',
        'Lead Unit Department', 'Lead Unit Contact Person', 'Lead Unit Contact Info',
        'Curriculum Officer', 'Number of Partners', 'Partner Agency Name',
        'Level', 'Category Age',
        'Nature of Partnership', 'Type of Extension', 'Type of Partnership',
        'Date Approved by Board', 'Duration', 'Inclusive Dates', 'Amount Involved',
        'SDG Number', 'Thematic Area', 'Extension Activity Period', 'Remarks'
    ]
    
    writer.writerow(['=== PARTNERSHIPS ==='])
    writer.writerow(partnership_headers)
    
    # Write partnership data
    for p in partnerships:
        row = [
            p.partship_id, p.moano or '', p.code or '', p.categ or '', p.datenotar or '',
            p.leadunit.department if p.leadunit else '', 
            p.leadunit.contper if p.leadunit else '',
            p.leadunit.continf if p.leadunit else '',
            p.curricoff or '', p.numpart or '', 
            p.partagency.nameagen if p.partagency else '',
            p.level or '', p.catage or '', p.natpart or '', p.tappext or '',
            p.tofpart or '', p.dappbor or '', p.duration or '', p.incldate or '',
            p.amntinvlv or '', p.sdgnum or '', p.themarea or '', p.extacperiod or '',
            p.remarks or ''
        ]
        writer.writerow(row)
    
    # Add separator and internal projects
    writer.writerow([])
    writer.writerow(['=== INTERNAL PROJECTS ==='])
    
    internal_headers = [
        'Internal Project ID', 'Partnership ID', 'Partnership Agency', 'Code',
        'Lead Unit Department', 'Collaborating Agency',
        'SDG Number', 'Thematic Area', 'Remarks'
    ]
    writer.writerow(internal_headers)
    
    # Write internal projects data
    internal_projects = InterFep.objects.select_related('lunitid', 'collabagenid', 'partship__partagency').all()
    partnerships_with_internal = set()
    
    for proj in internal_projects:
        row = [
            proj.interfepid, 
            proj.partship.partship_id if proj.partship else '',
            proj.partship.partagency.nameagen if proj.partship and proj.partship.partagency else '',
            proj.code or '',
            proj.lunitid.department if proj.lunitid else '',
            proj.collabagenid.nameagen if proj.collabagenid else '',
            '', proj.sdgnum or '',
            proj.themarea or '', proj.remarks or ''
        ]
        writer.writerow(row)
        if proj.partship:
            partnerships_with_internal.add(proj.partship.partship_id)
    
    # Add blank internal project forms
    internal_partnerships = Partnership.objects.filter(natpart__in=['Internal', 'Both'])
    for p in internal_partnerships:
        if p.partship_id not in partnerships_with_internal:
            row = [
                '', p.partship_id, 
                p.partagency.nameagen if p.partagency else '',
                '', p.leadunit.department if p.leadunit else '',
                p.partagency.nameagen if p.partagency else '',
                p.sdgnum or '', p.themarea or '', ''
            ]
            writer.writerow(row)
    
    return response

def populate_form_from_partnership(partnership, form_instance):
    """Populate form fields from partnership data"""
    if not partnership:
        return form_instance
    
    # Get auto-population data from partnership
    auto_data = partnership.get_auto_population_data()
    
    # Map partnership fields to form fields based on form type
    field_mappings = {
        'InterFepForm': {
            'lunitid': 'leadunit',
            'collabagenid': 'partagency',
            'curricoffid': 'curricoff',
        },
        'ExterFepForm': {
            'lunitid': 'leadunit',
            'collabgenid': 'partagency',
            'curricoffid': 'curricoff',
        },
        'AdvServicesForm': {
            'lunitid': 'leadunit',
            'collabagenciesid': 'partagency',
            'curricoff': 'curricoff',
        }
    }
    
    form_class_name = form_instance.__class__.__name__
    mappings = field_mappings.get(form_class_name, {})
    
    # Apply field mappings
    for form_field, partnership_field in mappings.items():
        if hasattr(form_instance, form_field) and partnership_field in auto_data:
            setattr(form_instance, form_field, auto_data[partnership_field])
    
    # Apply common fields
    common_fields = ['sdgnum', 'themarea', 'duration', 'incldate', 'amntinvlv']
    for field in common_fields:
        if hasattr(form_instance, field) and field in auto_data:
            setattr(form_instance, field, auto_data[field])
    
    return form_instance

def get_partnership_context(partnership_id):
    """Get comprehensive partnership context for forms"""
    try:
        partnership = Partnership.objects.select_related(
            'leadunit', 'partagency'
        ).get(pk=partnership_id)
        
        return {
            'partnership': partnership,
            'available_forms': partnership.get_available_form_types(),
            'auto_data': partnership.get_auto_population_data(),
            'form_status': get_form_creation_status(partnership)
        }
    except Partnership.DoesNotExist:
        return None

def get_form_creation_status(partnership):
    """Get status of form creation for partnership"""
    if not partnership:
        return {}
    
    status = {}
    available_types = partnership.get_available_form_types()
    
    from django.db.models import Count
    
    for form_type in available_types:
        if form_type == 'internal':
            count = InterFep.objects.filter(partship=partnership).count()
            status['internal'] = {'exists': count > 0, 'count': count}
        elif form_type == 'external':
            count = ExterFep.objects.filter(partship=partnership).count()
            status['external'] = {'exists': count > 0, 'count': count}
        elif form_type == 'advisory':
            count = AdvServices.objects.filter(partship=partnership).count()
            status['advisory'] = {'exists': count > 0, 'count': count}
    
    return status

def get_form_initial_data(partnership, form_type):
    """Get initial data for forms based on partnership and form type"""
    if not partnership:
        return {}
    
    # Get auto-population templates
    templates = AutoPopulationManager.auto_populate_forms(partnership, [form_type])
    
    if form_type in templates:
        return templates[form_type]
    
    return {}

def populate_sub_fields_from_partnership(partnership, form_initial):
    """Populate sub-field initial data from partnership"""
    if not partnership:
        return form_initial
    
    # Lead unit sub-fields
    if partnership.leadunit:
        form_initial.update({
            'lunit_department': partnership.leadunit.department or '',
            'lunit_contper': partnership.leadunit.contper or '',
            'lunit_continf': partnership.leadunit.continf or '',
        })
    
    # Partner agency sub-fields
    if partnership.partagency:
        form_initial.update({
            'collab_name': partnership.partagency.nameagen or '',
        })
    
    return form_initial

def prepare_form_initial_data(partnership, form_type):
    """Prepare initial data for form based on partnership and form type"""
    if not partnership:
        return {}
    
    # Get base auto-population data
    initial_data = get_form_initial_data(partnership, form_type)
    
    # Add sub-field data
    initial_data = populate_sub_fields_from_partnership(partnership, initial_data)
    
    return initial_data

def dashboard(request):
    """Main dashboard view"""
    context = {
        'partnership_count': Partnership.objects.count(),
        'internal_projects_count': InterFep.objects.count(),
        'external_projects_count': ExterFep.objects.count(),
        'advisory_services_count': AdvServices.objects.count(),
    }
    return render(request, 'dashboard.html', context)

# Partnership Views

class PartnershipFormView(FormView):
    """Partnership form view - create or edit"""
    template_name = 'partnership_form.html'
    form_class = PartnershipForm
    
    def form_valid(self, form):
        """Handle form validation and saving"""
        instance = form.save()
        
        # Handle files and images
        files = form.cleaned_data.get('files')
        images = form.cleaned_data.get('images')
        
        if files or images:
            handle_file_uploads(instance, files, images)
        
        return redirect('partnership_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['partnerships'] = Partnership.objects.all()
        return context

def _handle_partnership_sub_fields(request, form):
    """Helper function to handle partnership sub-field creation"""
    # Handle LeadUnit sub-fields
    leadunit_data = {
        'department': request.POST.get('leadunit_department'),
        'contper': request.POST.get('leadunit_contper'),
        'continf': request.POST.get('leadunit_continf'),
    }
    if any(leadunit_data.values()):
        try:
            leadunit, created = LeadUnit.objects.get_or_create(**leadunit_data)
            form.instance.leadunit = leadunit
        except (IntegrityError, ValidationError) as e:
            return f'Error creating lead unit: {escape(str(e))}'
    
    # Handle PartnerAgencies sub-fields
    partagency_data = {
        'nameagen': request.POST.get('partagency_name'),
    }
    if any(partagency_data.values()):
        try:
            partagency, created = PartnerAgencies.objects.get_or_create(**partagency_data)
            form.instance.partagency = partagency
        except IntegrityError:
            # If still fails, try to get existing record
            try:
                partagency = PartnerAgencies.objects.get(nameagen=partagency_data['nameagen'])
                form.instance.partagency = partagency
            except PartnerAgencies.DoesNotExist:
                return f'Error creating partner agency: {escape(partagency_data["nameagen"])}'
    
    return None

def _handle_partnership_file_uploads(request, partnership):
    """Helper function to handle partnership file uploads"""
    # Handle file uploads with security checks
    files = request.FILES.getlist('files')
    for file in files[:10]:
        # Validate file extension and size
        if file.size > 10 * 1024 * 1024:  # 10MB limit
            messages.warning(request, f'File {escape(file.name)} too large (max 10MB)')
            continue
        
        # Sanitize filename to prevent path traversal
        try:
            safe_filename = os_basename(file.name)
            # Additional validation - prevent path traversal
            if '..' in safe_filename or '/' in safe_filename or '\\' in safe_filename:
                messages.warning(request, f'Invalid filename: {escape(file.name)}')
                continue
            # Validate file extension
            allowed_extensions = ['.pdf', '.doc', '.docx', '.xls', '.xlsx', '.ppt', '.pptx', '.txt', '.csv', '.zip', '.rar', '.7z']
            if not any(safe_filename.lower().endswith(ext) for ext in allowed_extensions):
                messages.warning(request, f'File type not allowed: {escape(safe_filename)}')
                continue
            PartnershipFile.objects.create(
                partnership=partnership,
                file=file,
                filename=escape(safe_filename)
            )
        except (OSError, ValueError) as e:
            messages.error(request, f'File upload error: {escape(str(e))}')
            continue
    
    # Handle image uploads
    images = request.FILES.getlist('images')
    for image in images[:10]:
        try:
            safe_filename = os_basename(image.name)
            # Validate image extension
            allowed_image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.svg']
            if not any(safe_filename.lower().endswith(ext) for ext in allowed_image_extensions):
                messages.warning(request, f'Image type not allowed: {escape(safe_filename)}')
                continue
            PartnershipImage.objects.create(
                partnership=partnership,
                image=image,
                filename=escape(safe_filename)
            )
        except (OSError, ValueError) as e:
            messages.error(request, f'Image upload error: {escape(str(e))}')
            continue

def partnership_form(request, pk=None):
    """Partnership form view - create or edit"""
    partnership = get_object_or_404(Partnership, pk=pk) if pk else None
    
    # Get all partnerships for records table
    partnerships = Partnership.objects.select_related('leadunit', 'partagency').order_by('partship_id')
    
    if request.method == 'POST':
        form = PartnershipForm(request.POST, request.FILES, instance=partnership)
        if form.is_valid():
            print(f"DEBUG: Partnership form is valid")
            with transaction.atomic():
                # Handle sub-fields
                error_msg = _handle_partnership_sub_fields(request, form)
                if error_msg:
                    messages.error(request, error_msg)
                    return render(request, 'partnership_form.html', {'form': form})
                
                partnership = form.save()
                print(f"DEBUG: Partnership saved with ID: {partnership.partship_id}")
                    
                # Handle file uploads
                _handle_partnership_file_uploads(request, partnership)
                
                messages.success(request, 'Partnership saved successfully!')
                
                return redirect('partnership_form')
        else:
            print(f"DEBUG: Partnership form errors: {form.errors}")
            messages.error(request, f'Form validation failed: {form.errors}')
            context = {
                'form': form,
                'partnership': partnership,
                'partnerships': partnerships,
                'title': 'Edit Partnership' if partnership else 'Create Partnership'
            }
            return render(request, 'partnership_form.html', context)
    else:
        form = PartnershipForm(instance=partnership)
        
        # Auto-populate sub-fields if editing
        if partnership:
            # Set form field values directly for editing
            if hasattr(form, 'fields'):
                if 'moano' in form.fields:
                    form.fields['moano'].initial = partnership.moano
                if 'code' in form.fields:
                    form.fields['code'].initial = partnership.code
                if 'categ' in form.fields:
                    form.fields['categ'].initial = partnership.categ
                if 'datenotar' in form.fields:
                    form.fields['datenotar'].initial = partnership.datenotar
                if 'curricoff' in form.fields:
                    form.fields['curricoff'].initial = partnership.curricoff
                if 'numpart' in form.fields:
                    form.fields['numpart'].initial = partnership.numpart
                if 'level' in form.fields:
                    form.fields['level'].initial = partnership.level
                if 'catage' in form.fields:
                    form.fields['catage'].initial = partnership.catage
                if 'natpart' in form.fields:
                    form.fields['natpart'].initial = partnership.natpart
                if 'tappext' in form.fields:
                    form.fields['tappext'].initial = partnership.tappext
                if 'tofpart' in form.fields:
                    form.fields['tofpart'].initial = partnership.tofpart
                if 'dappbor' in form.fields:
                    form.fields['dappbor'].initial = partnership.dappbor
                if 'duration' in form.fields:
                    form.fields['duration'].initial = partnership.duration
                if 'incldate' in form.fields:
                    form.fields['incldate'].initial = partnership.incldate
                if 'amntinvlv' in form.fields:
                    form.fields['amntinvlv'].initial = partnership.amntinvlv
                if 'sdgnum' in form.fields:
                    form.fields['sdgnum'].initial = partnership.sdgnum
                if 'themarea' in form.fields:
                    form.fields['themarea'].initial = partnership.themarea
                if 'extacperiod' in form.fields:
                    form.fields['extacperiod'].initial = partnership.extacperiod
                if 'remarks' in form.fields:
                    form.fields['remarks'].initial = partnership.remarks
            
            form.initial.update({
                'leadunit_department': partnership.leadunit.department if partnership.leadunit else '',
                'leadunit_contper': partnership.leadunit.contper if partnership.leadunit else '',
                'leadunit_continf': partnership.leadunit.continf if partnership.leadunit else '',
                'partagency_name': partnership.partagency.nameagen if partnership.partagency else '',
            })
        
        # Add context about available forms if editing
        if partnership:
            available_forms = partnership.get_available_form_types()
            if available_forms:
                form_names = ', '.join([f.title() for f in available_forms])
                messages.info(
                    request,
                    f'This partnership can create: {form_names} forms. Use the navigation above to access them.'
                )
    
    context = {
        'form': form,
        'partnership': partnership,
        'partnerships': partnerships,
        'title': 'Edit Partnership' if partnership else 'Create Partnership'
    }
    return render(request, 'partnership_form.html', context)

def partnership_list(request):
    """Partnership list view"""
    partnerships = Partnership.objects.select_related('leadunit', 'partagency').all()
    paginator = Paginator(partnerships, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'title': 'Partnership Records'
    }
    return render(request, 'partnership_list.html', context)

class BaseRelatedFormView(FormView):
    """Base view for forms related to Partnership"""
    
    def get_initial(self):
        """
        Pre-populate form with partnership data if partnership_id provided
        """
        initial = super().get_initial()
        partnership_id = self.request.GET.get('partship_id')
        
        if partnership_id:
            try:
                partnership = Partnership.objects.get(partship_id=partnership_id)
                # Pre-populate form with partnership data
                initial.update({
                    'moano': partnership.moano,
                    'datenotar': partnership.datenotar,
                    'leadunit': partnership.leadunit,
                    'partagency': partnership.partagency,
                    'curricoff': partnership.curricoff,
                    'level': partnership.level,
                    'duration': partnership.duration,
                    'incldate': partnership.incldate,
                    'amntinvlv': partnership.amntinvlv,
                    'sdgnum': partnership.sdgnum,
                    'themarea': partnership.themarea
                })
            except Partnership.DoesNotExist:
                pass
        
        return initial

# Internal Projects Views
def _handle_internal_project_sub_fields(request, form):
    """Helper function to handle internal project sub-field creation"""
    # Handle projno field - ensure it's properly set
    projno_value = request.POST.get('projno')
    if projno_value:
        try:
            # Try to convert to integer, if it fails, treat as text
            try:
                projno_int = int(projno_value)
                projno_obj, created = ProjectNoInText.objects.get_or_create(
                    project_no=projno_int,
                    defaults={'proj_type': 'Internal'}
                )
            except ValueError:
                # If not a number, create with a default number and store text separately
                projno_obj, created = ProjectNoInText.objects.get_or_create(
                    project_no=0,  # Default number
                    defaults={'proj_type': 'Internal'}
                )
            form.instance.projno = projno_obj
        except Exception as e:
            return f'Error processing project number: {escape(str(e))}'
    
    # Handle sub-field data for LeadUnit
    leadunit_data = {
        'department': request.POST.get('lunit_department'),
        'contper': request.POST.get('lunit_contper'),
        'continf': request.POST.get('lunit_continf'),
    }
    if any(leadunit_data.values()):
        try:
            leadunit, created = LeadUnit.objects.get_or_create(**leadunit_data)
            form.instance.lunitid = leadunit
        except (IntegrityError, ValidationError) as e:
            return f'Error creating lead unit: {escape(str(e))}'

    # Handle sub-field data for Collaborating Agency
    collab_data = {
        'nameagen': request.POST.get('collab_name'),
    }
    if any(collab_data.values()):
        collab_agency, created = PartnerAgencies.objects.get_or_create(**collab_data)
        form.instance.collabagenid = collab_agency

    # Handle Community Survey data
    cmm_data = {
        'datecondt': request.POST.get('cmm_datecondt') or None,
        'dvstake': request.POST.get('cmm_dvstake') or None,
        'datepresent': request.POST.get('cmm_datepresent') or None,
        'datercvd': request.POST.get('cmm_datercvd') or None,
    }
    if any(cmm_data.values()):
        cmm_survey, created = CmmSrvy.objects.get_or_create(**cmm_data)
        form.instance.cmmsrvyid = cmm_survey

    # Handle Program data
    prog_apprvbudg = request.POST.get('prog_apprvbudg') or None
    prog_cntptfund = request.POST.get('prog_cntptfund') or None
    prog_totalbudg = None
    if prog_apprvbudg and prog_cntptfund:
        prog_totalbudg = int(prog_apprvbudg) + int(prog_cntptfund)
    elif prog_apprvbudg:
        prog_totalbudg = int(prog_apprvbudg)
    elif prog_cntptfund:
        prog_totalbudg = int(prog_cntptfund)
    
    prog_data = {
        'title': request.POST.get('prog_title'),
        'proglead': request.POST.get('prog_lead'),
        'apprvbudg': prog_apprvbudg,
        'cntptfund': prog_cntptfund,
        'totalbudg': prog_totalbudg,
        'incldateid': request.POST.get('prog_incldateid'),
        'moanoid': request.POST.get('prog_moanoid') or None,
    }
    if any(prog_data.values()):
        program, created = IfpoProg.objects.get_or_create(**prog_data)
        form.instance.ifpartofprogid = program

    # Handle Project data
    proj_apprvbudg = request.POST.get('proj_apprvbudg') or None
    proj_cntptfund = request.POST.get('proj_cntptfund') or None
    proj_totalbudg = None
    if proj_apprvbudg and proj_cntptfund:
        proj_totalbudg = int(proj_apprvbudg) + int(proj_cntptfund)
    elif proj_apprvbudg:
        proj_totalbudg = int(proj_apprvbudg)
    elif proj_cntptfund:
        proj_totalbudg = int(proj_cntptfund)
    
    proj_data = {
        'title': request.POST.get('proj_title'),
        'tmbrroles': request.POST.get('proj_tmbrroles'),
        'apprvbudg': proj_apprvbudg,
        'cntptfund': proj_cntptfund,
        'totalbudg': proj_totalbudg,
        'incldateid': request.POST.get('proj_incldateid'),
        'moanoid': request.POST.get('proj_moanoid') or None,
    }
    if any(proj_data.values()):
        project_obj, created = Proj.objects.get_or_create(**proj_data)
        form.instance.projid = project_obj

    # Handle Award data
    award_data = {
        'titleaward': request.POST.get('award_title'),
        'conferagn': request.POST.get('award_conferagn'),
        'date': request.POST.get('award_date') or None,
    }
    if any(award_data.values()):
        award, created = AwardRcvd.objects.get_or_create(**award_data)
        form.instance.awrdrcvdid = award
    
    return None

def _handle_internal_project_file_uploads(request, project):
    """Helper function to handle internal project file uploads"""
    # Handle file uploads with security validation
    files = request.FILES.getlist('files')
    file_objects = []
    for file in files[:10]:
        if file.size > FILE_SIZE_LIMIT:
            messages.warning(request, f'File {escape(file.name)} too large (max 10MB)')
            continue
        safe_filename = os_basename(file.name)
        if '..' in safe_filename or '/' in safe_filename or '\\' in safe_filename:
            messages.warning(request, f'Invalid filename: {escape(file.name)}')
            continue
        # Validate file extension
        allowed_extensions = ['.pdf', '.doc', '.docx', '.xls', '.xlsx', '.ppt', '.pptx', '.txt', '.csv', '.zip', '.rar', '.7z']
        if not any(safe_filename.lower().endswith(ext) for ext in allowed_extensions):
            messages.warning(request, f'File type not allowed: {escape(safe_filename)}')
            continue
        file_objects.append(InterFepFile(
            interfep=project,
            file=file,
            filename=escape(safe_filename)
        ))
    InterFepFile.objects.bulk_create(file_objects)
    
    # Handle image uploads
    images = request.FILES.getlist('images')
    image_objects = []
    for image in images[:10]:
        safe_filename = os_basename(image.name)
        # Validate image extension
        allowed_image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.svg']
        if not any(safe_filename.lower().endswith(ext) for ext in allowed_image_extensions):
            messages.warning(request, f'Image type not allowed: {escape(safe_filename)}')
            continue
        image_objects.append(InterFepImage(
            interfep=project,
            image=image,
            filename=escape(safe_filename)
        ))
    InterFepImage.objects.bulk_create(image_objects)

def internal_project_form(request, pk=None):
    """Internal project form view"""
    project = get_object_or_404(InterFep, pk=pk) if pk else None
    partnership_id = request.GET.get('partnership_id')
    
    # Get all partnerships that can create internal projects (ordered by ID)
    available_partnerships = Partnership.objects.filter(
        natpart__in=['Internal', 'Both']
    ).select_related('leadunit', 'partagency').order_by('partship_id')
    

    
    # Get recent internal projects for the records table
    internal_projects = InterFep.objects.select_related('lunitid', 'collabagenid', 'partship__partagency').all()[:20]

    # Auto-populate if partnership_id is provided
    initial_data = {}
    if partnership_id and not project:
        try:
            partnership = Partnership.objects.get(pk=partnership_id, natpart__in=['Internal', 'Both'])
            initial_data = {
                'partship': partnership,
                'lunitid': partnership.leadunit,
                'collabagenid': partnership.partagency,
                'curricoffid': partnership.curricoff,
                'sdgnum': partnership.sdgnum,
                'themarea': partnership.themarea,
            }
        except Partnership.DoesNotExist:
            pass
    
    if request.method == 'POST':
        form = InterFepForm(request.POST, request.FILES, instance=project, partnership_id=partnership_id)
        if form.is_valid():
            print(f"DEBUG: Internal form is valid")
            with transaction.atomic():
                # Set the partnership if provided
                partship = form.cleaned_data.get('partship')
                if partship:
                    form.instance.partship = partship
                    print(f"DEBUG: Partnership set to: {partship.partship_id}")
                else:
                    print("DEBUG: No partnership selected - proceeding without partnership")
                
                # Handle sub-fields
                error_msg = _handle_internal_project_sub_fields(request, form)
                if error_msg:
                    messages.error(request, error_msg)
                    context = {
                        'form': form,
                        'project': project,
                        'partnership_id': partnership_id,
                        'available_partnerships': available_partnerships,
                        'internal_projects': internal_projects,
                        'title': 'Edit Internal Project' if project else 'Create Internal Project'
                    }
                    return render(request, 'internal_project_form.html', context)
                
                project = form.save()
                print(f"DEBUG: Internal project saved with ID: {project.interfepid}")
                
                # Handle file uploads
                _handle_internal_project_file_uploads(request, project)
                
                messages.success(request, 'Internal project saved successfully!')
                return redirect('internal_project_form')
        else:
            print(f"DEBUG: Internal form errors: {form.errors}")
            messages.error(request, f'Form validation failed: {form.errors}')
            context = {
                'form': form,
                'project': project,
                'partnership_id': partnership_id,
                'available_partnerships': available_partnerships,
                'internal_projects': internal_projects,
                'title': 'Edit Internal Project' if project else 'Create Internal Project'
            }
            return render(request, 'internal_project_form.html', context)
    else:
        form = InterFepForm(instance=project, partnership_id=partnership_id, initial=initial_data)
        
        # Auto-populate sub-fields if partnership_id provided
        if partnership_id and not project:
            try:
                partnership = Partnership.objects.select_related('leadunit', 'partagency').get(pk=partnership_id)
                form.initial.update({
                    'lunit_department': partnership.leadunit.department if partnership.leadunit else '',
                    'lunit_contper': partnership.leadunit.contper if partnership.leadunit else '',
                    'lunit_continf': partnership.leadunit.continf if partnership.leadunit else '',
                    'collab_name': partnership.partagency.nameagen if partnership.partagency else '',
                })
            except Partnership.DoesNotExist:
                pass

        # Auto-populate sub-fields if editing
        if project:
            # Set main form field values that exist in the form
            if hasattr(form, 'fields'):
                if 'projno' in form.fields:
                    form.fields['projno'].initial = project.projno.project_no if project.projno else ''
                if 'code' in form.fields:
                    form.fields['code'].initial = project.code
                if 'categid' in form.fields:
                    form.fields['categid'].initial = project.categid
                if 'curricoffid' in form.fields:
                    form.fields['curricoffid'].initial = project.curricoffid
                if 'dateapprec' in form.fields:
                    form.fields['dateapprec'].initial = project.dateapprec
                if 'dateappborid' in form.fields:
                    form.fields['dateappborid'].initial = project.dateappborid
                if 'dateincep' in form.fields:
                    form.fields['dateincep'].initial = project.dateincep
                if 'benef' in form.fields:
                    form.fields['benef'].initial = project.benef
                if 'sdgnum' in form.fields:
                    form.fields['sdgnum'].initial = project.sdgnum
                if 'themarea' in form.fields:
                    form.fields['themarea'].initial = project.themarea
                if 'remarks' in form.fields:
                    form.fields['remarks'].initial = project.remarks
            
            form.initial.update({
                'lunit_department': project.lunitid.department if project.lunitid else '',
                'lunit_contper': project.lunitid.contper if project.lunitid else '',
                'lunit_continf': project.lunitid.continf if project.lunitid else '',
                'collab_name': project.collabagenid.nameagen if project.collabagenid else '',
                'cmm_datecondt': project.cmmsrvyid.datecondt if project.cmmsrvyid else '',
                'cmm_dvstake': project.cmmsrvyid.dvstake if project.cmmsrvyid else '',
                'cmm_datepresent': project.cmmsrvyid.datepresent if project.cmmsrvyid else '',
                'cmm_datercvd': project.cmmsrvyid.datercvd if project.cmmsrvyid else '',
                'prog_title': project.ifpartofprogid.title if project.ifpartofprogid else '',
                'prog_lead': project.ifpartofprogid.proglead if project.ifpartofprogid else '',
                'prog_apprvbudg': project.ifpartofprogid.apprvbudg if project.ifpartofprogid else '',
                'prog_cntptfund': project.ifpartofprogid.cntptfund if project.ifpartofprogid else '',
                'prog_totalbudg': project.ifpartofprogid.totalbudg if project.ifpartofprogid else '',
                'prog_incldateid': project.ifpartofprogid.incldateid if project.ifpartofprogid else '',
                'proj_title': project.projid.title if project.projid else '',
                'proj_tmbrroles': project.projid.tmbrroles if project.projid else '',
                'proj_apprvbudg': project.projid.apprvbudg if project.projid else '',
                'proj_cntptfund': project.projid.cntptfund if project.projid else '',
                'proj_totalbudg': project.projid.totalbudg if project.projid else '',
                'proj_incldateid': project.projid.incldateid if project.projid else '',
                'award_title': project.awrdrcvdid.titleaward if project.awrdrcvdid else '',
                'award_conferagn': project.awrdrcvdid.conferagn if project.awrdrcvdid else '',
                'award_date': project.awrdrcvdid.date if project.awrdrcvdid else '',
            })

    context = {
        'form': form,
        'project': project,
        'partnership_id': partnership_id,
        'available_partnerships': available_partnerships,
        'internal_projects': internal_projects,
        'title': 'Edit Internal Project' if project else 'Create Internal Project'
    }
    return render(request, 'internal_project_form.html', context)

def internal_project_list(request):
    """Internal projects list view"""
    projects = InterFep.objects.select_related('lunitid', 'collabagenid', 'partship').all()
    paginator = Paginator(projects, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'title': 'Internal Project Records'
    }
    return render(request, 'internal_project_list.html', context)

def external_project_list(request):
    """External projects list view"""
    projects = ExterFep.objects.select_related('lunitid', 'collabgenid', 'partship').all()
    paginator = Paginator(projects, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'title': 'External Project Records'
    }
    return render(request, 'external_project_list.html', context)

# External Projects Views
def external_project_form(request, pk=None):
    """External project form view"""
    project = get_object_or_404(ExterFep, pk=pk) if pk else None
    partnership_id = request.GET.get('partnership_id')
    
    # Get all partnerships that can create external projects (ordered by ID)
    available_partnerships = Partnership.objects.filter(
        natpart__in=['External', 'Both']
    ).select_related('leadunit', 'partagency').order_by('partship_id')

    # Get recent external projects for the records table
    external_projects = ExterFep.objects.select_related('partship').all()[:20]
    
    
    # Auto-populate if partnership_id is provided
    initial_data = {}
    if partnership_id and not project:
        try:
            partnership = Partnership.objects.get(pk=partnership_id, natpart__in=['External', 'Both'])
            initial_data = {
                'partship': partnership,
                'lunitid': partnership.leadunit,
                'collabgenid': partnership.partagency,
                'curricoffid': partnership.curricoff,
                'sdgnum': partnership.sdgnum,
                'themarea': partnership.themarea,
            }
        except Partnership.DoesNotExist:
            pass
    
    if request.method == 'POST':
        form = ExterFepForm(request.POST, request.FILES, instance=project, partnership_id=partnership_id)
        if form.is_valid():
            print(f"DEBUG: External form is valid")
            with transaction.atomic():
                # Set the partnership if provided
                partship = form.cleaned_data.get('partship')
                if partship:
                    form.instance.partship = partship
                    print(f"DEBUG: Partnership set to: {partship.partship_id}")
                else:
                    print("DEBUG: No partnership selected - proceeding without partnership")
                
                # Handle projno field - ensure it's properly set
                projno_value = request.POST.get('projno')
                if projno_value:
                    try:
                        # Try to convert to integer, if it fails, treat as text
                        try:
                            projno_int = int(projno_value)
                            projno_obj, created = ProjectNoInText.objects.get_or_create(
                                project_no=projno_int,
                                defaults={'proj_type': 'External'}
                            )
                        except ValueError:
                            # If not a number, create with a default number and store text separately
                            projno_obj, created = ProjectNoInText.objects.get_or_create(
                                project_no=0,  # Default number
                                defaults={'proj_type': 'External'}
                            )
                        form.instance.projno = projno_obj
                    except Exception as e:
                        messages.error(request, f'Error processing project number: {escape(str(e))}')
                        context = {
                            'form': form,
                            'project': project,
                            'partnership_id': partnership_id,
                            'available_partnerships': available_partnerships,
                            'external_projects': external_projects,
                            'title': 'Edit External Project' if project else 'Create External Project'
                        }
                        return render(request, 'external_project_form.html', context)
                
                # Handle sub-field data for LeadUnit
                leadunit_data = {
                    'department': request.POST.get('lunit_department'),
                    'contper': request.POST.get('lunit_contper'),
                    'continf': request.POST.get('lunit_continf'),
                }
                if any(leadunit_data.values()):
                    leadunit, created = LeadUnit.objects.get_or_create(**leadunit_data)
                    form.instance.lunitid = leadunit

                # Handle sub-field data for Collaborating Agency
                collab_data = {
                    'nameagen': request.POST.get('collab_name'),
                }
                if any(collab_data.values()):
                    collab_agency, created = PartnerAgencies.objects.get_or_create(**collab_data)
                    form.instance.collabgenid = collab_agency

                # Handle External Program data
                extprog_abfund = request.POST.get('extprog_abfundagency') or None
                extprog_cpfund = request.POST.get('extprog_cpfundcvsu') or None
                extprog_total = None
                if extprog_abfund and extprog_cpfund:
                    extprog_total = int(extprog_abfund) + int(extprog_cpfund)
                elif extprog_abfund:
                    extprog_total = int(extprog_abfund)
                elif extprog_cpfund:
                    extprog_total = int(extprog_cpfund)
                
                extprog_data = {
                    'title': request.POST.get('extprog_title'),
                    'proglead': request.POST.get('extprog_proglead'),
                    'abfundagency': extprog_abfund,
                    'cpfundcvsu': extprog_cpfund,
                    'tbudget': extprog_total,
                    'incldates': request.POST.get('extprog_incldates'),
                    'moano_id': request.POST.get('extprog_moano_id') or None,
                }
                if any(extprog_data.values()):
                    extprog, created = IfpoProgExter.objects.get_or_create(**extprog_data)
                    form.instance.ifpoprogid = extprog

                # Handle External Project data
                extproj_abfund = request.POST.get('extproj_abfundagency') or None
                extproj_cpfund = request.POST.get('extproj_cpfundcvsu') or None
                extproj_total = None
                if extproj_abfund and extproj_cpfund:
                    extproj_total = int(extproj_abfund) + int(extproj_cpfund)
                elif extproj_abfund:
                    extproj_total = int(extproj_abfund)
                elif extproj_cpfund:
                    extproj_total = int(extproj_cpfund)
                
                extproj_data = {
                    'title': request.POST.get('extproj_title'),
                    'tmembers': request.POST.get('extproj_tmembers'),
                    'abfundagency': extproj_abfund,
                    'cpfundcvsu': extproj_cpfund,
                    'tbudget': extproj_total,
                    'incldates': request.POST.get('extproj_incldates'),
                    'moano_id': request.POST.get('extproj_moano_id') or None,
                }
                if any(extproj_data.values()):
                    extproj, created = ProjExter.objects.get_or_create(**extproj_data)
                    form.instance.projid = extproj

                # Handle External Award data
                extaward_data = {
                    'toaward': request.POST.get('extaward_toaward'),
                    'conferagency': request.POST.get('extaward_conferagency'),
                    'date': request.POST.get('extaward_date'),
                }
                if any(extaward_data.values()):
                    extaward, created = AwardsRcvdExter.objects.get_or_create(**extaward_data)
                    form.instance.awardrcvdid = extaward

                project = form.save()
                print(f"DEBUG: External project saved with ID: {project.exterfep_id}")
                
                # Handle file uploads with security validation
                files = request.FILES.getlist('files')
                for file in files[:10]:
                    if file.size > FILE_SIZE_LIMIT:
                        messages.warning(request, f'File {escape(file.name)} too large (max 10MB)')
                        continue
                    safe_filename = os_basename(file.name)
                    if '..' in safe_filename or '/' in safe_filename or '\\' in safe_filename:
                        messages.warning(request, f'Invalid filename: {escape(file.name)}')
                        continue
                    # Validate file extension
                    allowed_extensions = ['.pdf', '.doc', '.docx', '.xls', '.xlsx', '.ppt', '.pptx', '.txt', '.csv', '.zip', '.rar', '.7z']
                    if not any(safe_filename.lower().endswith(ext) for ext in allowed_extensions):
                        messages.warning(request, f'File type not allowed: {escape(safe_filename)}')
                        continue
                    ExterFepFile.objects.create(
                        exterfep=project,
                        file=file,
                        filename=escape(safe_filename)
                    )
                
                # Handle image uploads with security validation
                images = request.FILES.getlist('images')
                for image in images[:10]:
                    try:
                        safe_filename = os_basename(image.name)
                        # Validate image extension
                        allowed_image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.svg']
                        if not any(safe_filename.lower().endswith(ext) for ext in allowed_image_extensions):
                            messages.warning(request, f'Image type not allowed: {escape(safe_filename)}')
                            continue
                        ExterFepImage.objects.create(
                            exterfep=project,
                            image=image,
                            filename=escape(safe_filename)
                        )
                    except (OSError, ValueError) as e:
                        messages.error(request, f'Image upload error: {escape(str(e))}')
                        continue
                
                messages.success(request, 'External project saved successfully!')
                return redirect('external_project_form')
        else:
            print(f"DEBUG: External form errors: {form.errors}")
            messages.error(request, f'Form validation failed: {form.errors}')
            context = {
                'form': form,
                'project': project,
                'partnership_id': partnership_id,
                'available_partnerships': available_partnerships,
                'external_projects': external_projects,
                'title': 'Edit External Project' if project else 'Create External Project'
            }
            return render(request, 'external_project_form.html', context)
    else:
        form = ExterFepForm(instance=project, partnership_id=partnership_id, initial=initial_data)

        # Auto-populate sub-fields if partnership_id provided for external projects
        if partnership_id and not project:
            try:
                partnership = Partnership.objects.select_related('leadunit', 'partagency').get(pk=partnership_id)
                form.initial.update({
                    'lunit_department': partnership.leadunit.department if partnership.leadunit else '',
                    'lunit_contper': partnership.leadunit.contper if partnership.leadunit else '',
                    'lunit_continf': partnership.leadunit.continf if partnership.leadunit else '',
                    'collab_name': partnership.partagency.nameagen if partnership.partagency else '',
                })
            except Partnership.DoesNotExist:
                pass

        # Auto-populate sub-fields if editing
        if project:
            # Set main form field values that exist in the form
            if hasattr(form, 'fields'):
                if 'projno' in form.fields:
                    form.fields['projno'].initial = project.projno.project_no if project.projno else ''
                if 'code' in form.fields:
                    form.fields['code'].initial = project.code
                if 'categid' in form.fields:
                    form.fields['categid'].initial = project.categid
                if 'curricoffid' in form.fields:
                    form.fields['curricoffid'].initial = project.curricoffid
                if 'fundagency' in form.fields:
                    form.fields['fundagency'].initial = project.fundagency
                if 'dtappfagency' in form.fields:
                    form.fields['dtappfagency'].initial = project.dtappfagency
                if 'dtinmeet' in form.fields:
                    form.fields['dtinmeet'].initial = project.dtinmeet
                if 'benef' in form.fields:
                    form.fields['benef'].initial = project.benef
                if 'sdgnum' in form.fields:
                    form.fields['sdgnum'].initial = project.sdgnum
                if 'themarea' in form.fields:
                    form.fields['themarea'].initial = project.themarea
                if 'remarks' in form.fields:
                    form.fields['remarks'].initial = project.remarks
            
            form.initial.update({
                'lunit_department': project.lunitid.department if project.lunitid else '',
                'lunit_contper': project.lunitid.contper if project.lunitid else '',
                'lunit_continf': project.lunitid.continf if project.lunitid else '',
                'collab_name': project.collabgenid.nameagen if project.collabgenid else '',
                'extprog_title': project.ifpoprogid.title if project.ifpoprogid else '',
                'extprog_proglead': project.ifpoprogid.proglead if project.ifpoprogid else '',
                'extprog_abfundagency': project.ifpoprogid.abfundagency if project.ifpoprogid else '',
                'extprog_cpfundcvsu': project.ifpoprogid.cpfundcvsu if project.ifpoprogid else '',
                'extprog_tbudget': project.ifpoprogid.tbudget if project.ifpoprogid else '',
                'extprog_incldates': project.ifpoprogid.incldates if project.ifpoprogid else '',
                'extproj_title': project.projid.title if project.projid else '',
                'extproj_tmembers': project.projid.tmembers if project.projid else '',
                'extproj_abfundagency': project.projid.abfundagency if project.projid else '',
                'extproj_cpfundcvsu': project.projid.cpfundcvsu if project.projid else '',
                'extproj_tbudget': project.projid.tbudget if project.projid else '',
                'extproj_incldates': project.projid.incldates if project.projid else '',
                'extaward_toaward': project.awardrcvdid.toaward if project.awardrcvdid else '',
                'extaward_conferagency': project.awardrcvdid.conferagency if project.awardrcvdid else '',
                'extaward_date': project.awardrcvdid.date if project.awardrcvdid else '',
            })
    
    context = {
        'form': form,
        'project': project,
        'partnership_id': partnership_id,
        'available_partnerships': available_partnerships,
        'external_projects': external_projects,
        'title': 'Edit External Project' if project else 'Create External Project'
    }
    return render(request, 'external_project_form.html', context)

# Advisory Services Views
def advisory_services_form(request, pk=None):
    """Advisory services form view"""
    service = get_object_or_404(AdvServices, pk=pk) if pk else None
    partnership_id = request.GET.get('partnership_id')
    
    # Get all partnerships that can create advisory services (ordered by ID)
    available_partnerships = Partnership.objects.filter(
        natpart='Advisory'
    ).select_related('leadunit', 'partagency').order_by('partship_id')
    
     # Get all advisory services for the records table
    advisory_services = AdvServices.objects.select_related('lunitid', 'collabagenciesid', 'partship', 'clientsid', 'adservoverall', 'adservtimeliness', 'estexsourcefund').order_by('advservices_id')
    

    # Auto-populate if partnership_id is provided
    initial_data = {}
    if partnership_id and not service:
        try:
            partnership = Partnership.objects.get(pk=partnership_id, natpart='Advisory')
            initial_data = {
                'partship': partnership,
                'lunitid': partnership.leadunit,
                'collabagenciesid': partnership.partagency,
                'curricoff': partnership.curricoff,
                'sdgnum': partnership.sdgnum,
                'themarea': partnership.themarea,
            }
        except Partnership.DoesNotExist:
            pass
    
    if request.method == 'POST':
        form = AdvServicesForm(request.POST, request.FILES, instance=service, partnership_id=partnership_id)
        if form.is_valid():
            print(f"DEBUG: Advisory form is valid")
            with transaction.atomic():
                # Set the partnership if provided
                partship = form.cleaned_data.get('partship')
                if partship:
                    form.instance.partship = partship
                    print(f"DEBUG: Partnership set to: {partship.partship_id}")
                else:
                    print("DEBUG: No partnership selected - proceeding without partnership")
                 # Handle sub-field data for LeadUnit
                leadunit_data = {
                    'department': request.POST.get('lunit_department'),
                    'contper': request.POST.get('lunit_contper'),
                    'continf': request.POST.get('lunit_continf'),
                }
                if any(leadunit_data.values()):
                    leadunit, created = LeadUnit.objects.get_or_create(**leadunit_data)
                    form.instance.lunitid = leadunit

                # Handle sub-field data for Collaborating Agency
                collab_data = {
                    'nameagen': request.POST.get('collab_name'),
                }
                if any(collab_data.values()):
                    try:
                        collab_agency, created = PartnerAgencies.objects.get_or_create(**collab_data)
                        form.instance.collabagenciesid = collab_agency
                    except IntegrityError:
                        try:
                            collab_agency = PartnerAgencies.objects.get(nameagen=collab_data['nameagen'])
                            form.instance.collabagenciesid = collab_agency
                        except PartnerAgencies.DoesNotExist:
                            messages.error(request, f'Error creating collaborating agency: {escape(collab_data["nameagen"])}')
                            return render(request, 'advisory_services_form.html', {'form': form})

                # Handle Client data
                client_data = {
                    'name': request.POST.get('client_name'),
                    'organization': request.POST.get('client_organization'),
                    'sex': request.POST.get('client_sex'),
                    'categ': request.POST.get('client_categ'),
                    'number': request.POST.get('client_number'),
                }
                if any(client_data.values()):
                    client, created = AdvSClients.objects.get_or_create(**client_data)
                    form.instance.clientsid = client

                # Handle Overall Rating data
                overall_data = {
                    'one': request.POST.get('overall_one') or None,
                    'two': request.POST.get('overall_two') or None,
                    'three': request.POST.get('overall_three') or None,
                    'four': request.POST.get('overall_four') or None,
                    'five': request.POST.get('overall_five') or None,
                }
                if any(overall_data.values()):
                    overall_rating, created = AsOverallRating.objects.get_or_create(**overall_data)
                    form.instance.adservoverall = overall_rating

                # Handle Timeliness Rating data
                timeliness_data = {
                    'one': request.POST.get('timeliness_one') or None,
                    'two': request.POST.get('timeliness_two') or None,
                    'three': request.POST.get('timeliness_three') or None,
                    'four': request.POST.get('timeliness_four') or None,
                    'five': request.POST.get('timeliness_five') or None,
                }
                if any(timeliness_data.values()):
                    timeliness_rating, created = AsTimeliness.objects.get_or_create(**timeliness_data)
                    form.instance.adservtimeliness = timeliness_rating

                # Handle External Source Fund data
                fund_data = {
                    'amntchcvsu': request.POST.get('fund_amntchcvsu') or None,
                    'amntchpartagency': request.POST.get('fund_amntchpartagency') or None,
                    'partagencyname': request.POST.get('fund_partagencyname'),
                }
                if any(fund_data.values()):
                    fund, created = EstExsFund.objects.get_or_create(**fund_data)
                    form.instance.estexsourcefund = fund



                service = form.save()
                print(f"DEBUG: Advisory service saved with ID: {service.advservices_id}")
                
                # Handle file uploads with security validation
                files = request.FILES.getlist('files')
                for file in files[:10]:
                    if file.size > FILE_SIZE_LIMIT:
                        messages.warning(request, f'File {escape(file.name)} too large (max 10MB)')
                        continue
                    safe_filename = os_basename(file.name)
                    if '..' in safe_filename or '/' in safe_filename or '\\' in safe_filename:
                        messages.warning(request, f'Invalid filename: {escape(file.name)}')
                        continue
                    # Validate file extension
                    allowed_extensions = ['.pdf', '.doc', '.docx', '.xls', '.xlsx', '.ppt', '.pptx', '.txt', '.csv', '.zip', '.rar', '.7z']
                    if not any(safe_filename.lower().endswith(ext) for ext in allowed_extensions):
                        messages.warning(request, f'File type not allowed: {escape(safe_filename)}')
                        continue
                    AdvServicesFile.objects.create(
                        advservices=service,
                        file=file,
                        filename=escape(safe_filename)
                    )
                
                # Handle image uploads with security validation
                images = request.FILES.getlist('images')
                for image in images[:10]:
                    try:
                        safe_filename = os_basename(image.name)
                        # Validate image extension
                        allowed_image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.svg']
                        if not any(safe_filename.lower().endswith(ext) for ext in allowed_image_extensions):
                            messages.warning(request, f'Image type not allowed: {escape(safe_filename)}')
                            continue
                        AdvServicesImage.objects.create(
                            advservices=service,
                            image=image,
                            filename=escape(safe_filename)
                        )
                    except (OSError, ValueError) as e:
                        messages.error(request, f'Image upload error: {escape(str(e))}')
                        continue
                
                messages.success(request, 'Advisory service saved successfully!')
                return redirect('advisory_services_form')
        else:
            print(f"DEBUG: Advisory form errors: {form.errors}")
            messages.error(request, f'Form validation failed: {form.errors}')
            context = {
                'form': form,
                'service': service,
                'partnership_id': partnership_id,
                'available_partnerships': available_partnerships,
                'advisory_services': advisory_services,
                'title': 'Edit Advisory Service' if service else 'Create Advisory Service'
            }
            return render(request, 'advisory_services_form.html', context)
    else:
        form = AdvServicesForm(instance=service, partnership_id=partnership_id, initial=initial_data)
    
        # Auto-populate sub-fields if partnership_id provided for advisory services
        if partnership_id and not service:
            try:
                partnership = Partnership.objects.select_related('leadunit', 'partagency').get(pk=partnership_id)
                form.initial.update({
                    'lunit_department': partnership.leadunit.department if partnership.leadunit else '',
                    'lunit_contper': partnership.leadunit.contper if partnership.leadunit else '',
                    'lunit_continf': partnership.leadunit.continf if partnership.leadunit else '',
                    'collab_name': partnership.partagency.nameagen if partnership.partagency else '',
                })
            except Partnership.DoesNotExist:
                pass

        # Auto-populate sub-fields if editing
        if service:
            # Set main form field values that exist in the form
            if hasattr(form, 'fields'):
                if 'number' in form.fields:
                    form.fields['number'].initial = service.number
                if 'code' in form.fields:
                    form.fields['code'].initial = service.code
                if 'curricoff' in form.fields:
                    form.fields['curricoff'].initial = service.curricoff
                if 'adservprov' in form.fields:
                    form.fields['adservprov'].initial = service.adservprov
                if 'incldates' in form.fields:
                    form.fields['incldates'].initial = service.incldates
                if 'venue' in form.fields:
                    form.fields['venue'].initial = service.venue
                if 'totaladservreq' in form.fields:
                    form.fields['totaladservreq'].initial = service.totaladservreq
                if 'totaladservreqrespond' in form.fields:
                    form.fields['totaladservreqrespond'].initial = service.totaladservreqrespond
                if 'sdgnum' in form.fields:
                    form.fields['sdgnum'].initial = service.sdgnum
                if 'themarea' in form.fields:
                    form.fields['themarea'].initial = service.themarea
                if 'remarks' in form.fields:
                    form.fields['remarks'].initial = service.remarks
            
            form.initial.update({
                'lunit_department': service.lunitid.department if service.lunitid else '',
                'lunit_contper': service.lunitid.contper if service.lunitid else '',
                'lunit_continf': service.lunitid.continf if service.lunitid else '',
                'collab_name': service.collabagenciesid.nameagen if service.collabagenciesid else '',
                'client_name': service.clientsid.name if service.clientsid else '',
                'client_organization': service.clientsid.organization if service.clientsid else '',
                'client_sex': service.clientsid.sex if service.clientsid else '',
                'client_categ': service.clientsid.categ if service.clientsid else '',
                'client_number': service.clientsid.number if service.clientsid else '',
                'overall_one': service.adservoverall.one if service.adservoverall else '',
                'overall_two': service.adservoverall.two if service.adservoverall else '',
                'overall_three': service.adservoverall.three if service.adservoverall else '',
                'overall_four': service.adservoverall.four if service.adservoverall else '',
                'overall_five': service.adservoverall.five if service.adservoverall else '',
                'timeliness_one': service.adservtimeliness.one if service.adservtimeliness else '',
                'timeliness_two': service.adservtimeliness.two if service.adservtimeliness else '',
                'timeliness_three': service.adservtimeliness.three if service.adservtimeliness else '',
                'timeliness_four': service.adservtimeliness.four if service.adservtimeliness else '',
                'timeliness_five': service.adservtimeliness.five if service.adservtimeliness else '',
                'fund_amntchcvsu': service.estexsourcefund.amntchcvsu if service.estexsourcefund else '',
                'fund_amntchpartagency': service.estexsourcefund.amntchpartagency if service.estexsourcefund else '',
                'fund_partagencyname': service.estexsourcefund.partagencyname if service.estexsourcefund else '',
            })
    
    context = {
        'form': form,
        'service': service,
        'partnership_id': partnership_id,
        'available_partnerships': available_partnerships,
        'advisory_services': advisory_services,
        'title': 'Edit Advisory Service' if service else 'Create Advisory Service'
    }
    return render(request, 'advisory_services_form.html', context)

def advisory_services_list(request):
    """Advisory services list view"""
    services = AdvServices.objects.select_related('partship').all()
    paginator = Paginator(services, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'title': 'Advisory Service Records'
    }
    return render(request, 'advisory_services_list.html', context)

# Navigation and utility views
def select_form_type(request, partnership_id):
    """View to select which form to create based on partnership nature"""
    try:
        partnership = Partnership.objects.get(pk=partnership_id)
        
        available_forms = []
        if partnership.natpart in ['Internal', 'Both']:
            available_forms.append({
                'name': 'Internal Project',
                'url': reverse('internal_project_form') + f'?partnership_id={partnership_id}',
                'description': 'Create an internally funded project'
            })
        
        if partnership.natpart in ['External', 'Both']:
            available_forms.append({
                'name': 'External Project', 
                'url': reverse('external_project_form') + f'?partnership_id={partnership_id}',
                'description': 'Create an externally funded project'
            })
        
        if partnership.natpart == 'Advisory':
            available_forms.append({
                'name': 'Advisory Service',
                'url': reverse('advisory_services_form') + f'?partnership_id={partnership_id}',
                'description': 'Create an advisory service'
            })
        
        context = {
            'partnership': partnership,
            'available_forms': available_forms,
            'title': 'Select Form Type'
        }
        return render(request, 'select_form_type.html', context)
        
    except Partnership.DoesNotExist:
        messages.error(request, 'Partnership not found')
        return redirect('partnership_list')

@require_http_methods(["GET"])
def get_partnerships_by_nature(request):
    """API endpoint to get partnerships filtered by nature"""
    nature = request.GET.get('nature')
    partnerships = Partnership.objects.select_related('partagency').filter(
        natpart__icontains=nature
    )[:100] if nature else Partnership.objects.select_related('partagency').all()[:100]
    
    data = [{
        'id': p.partship_id,
        'display_name': f"Partnership {p.partship_id} - {p.partagency.nameagen if p.partagency else 'No Agency'}"
    } for p in partnerships]
    
    return JsonResponse({'partnerships': data})

@require_http_methods(["GET"])
def get_partnership_data(request, partnership_id):
    """API endpoint to get partnership data by ID for auto-population"""
    try:
        partnership = Partnership.objects.select_related('leadunit', 'partagency').get(pk=partnership_id)
        
        # Base data for all forms (Internal, External, Advisory)
        data = {
            'id': partnership.partship_id,
            'natpart': partnership.natpart,
            'code': partnership.code or '',
            'curricoff': partnership.curricoff or '',
            'sdgnum': partnership.sdgnum or '',
            'themarea': partnership.themarea or '',
            'remarks': partnership.remarks or '',
            # Lead unit data - Enhanced null handling
            'lunit_department': partnership.leadunit.department if partnership.leadunit and partnership.leadunit.department else '',
            'lunit_contper': partnership.leadunit.contper if partnership.leadunit and partnership.leadunit.contper else '',
            'lunit_continf': partnership.leadunit.continf if partnership.leadunit and partnership.leadunit.continf else '',
        }
        
        # Additional data for Internal and External forms only (not Advisory)
        if partnership.natpart in ['Internal', 'External', 'Both']:
            data.update({
                # Partner agency name for Internal/External forms
                'collab_name': partnership.partagency.nameagen if partnership.partagency else '',
            })
        
        return JsonResponse(data)
    except Partnership.DoesNotExist:
        return JsonResponse({'error': 'Partnership not found'}, status=404)

# Add API endpoint for partnership data
@require_http_methods(["GET"])
def api_partnership_data(request, partnership_id):
    """API endpoint for JavaScript auto-population"""
    return get_partnership_data(request, partnership_id)

# Export Views
def export_partnerships(request):
    """Export partnerships to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="partnerships.csv"'
    
    writer = csv.writer(response)
    partnerships = Partnership.objects.select_related('leadunit', 'partagency').order_by('partship_id')
    
    headers = ['Partnership ID', 'MOA Number', 'Code', 'Category', 'Date Notarized',
               'Lead Unit Department', 'Lead Unit Contact Person', 'Lead Unit Contact Info',
               'Curriculum Officer', 'Number of Partners', 'Partner Agency Name',
               'Level', 'Category Age',
               'Nature of Partnership', 'Type of Extension', 'Type of Partnership',
               'Date Approved by Board', 'Duration', 'Inclusive Dates', 'Amount Involved',
               'SDG Number', 'Thematic Area', 'Extension Activity Period', 'Remarks']
    writer.writerow(headers)
    
    for p in partnerships:
        writer.writerow([
            p.partship_id, p.moano or '', p.code or '', p.categ or '', p.datenotar or '',
            p.leadunit.department if p.leadunit else '', 
            p.leadunit.contper if p.leadunit else '',
            p.leadunit.continf if p.leadunit else '',
            p.curricoff or '', p.numpart or '', 
            p.partagency.nameagen if p.partagency else '',
            p.partagency.headagen if p.partagency else '',
            p.partagency.contdet if p.partagency else '',
            p.level or '', p.catage or '', p.natpart or '', p.tappext or '',
            p.tofpart or '', p.dappbor or '', p.duration or '', p.incldate or '',
            p.amntinvlv or '', p.sdgnum or '', p.themarea or '', p.extacperiod or '',
            p.remarks or ''
        ])
    
    return response

def export_internal_projects(request):
    """Export internal projects to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="internal_projects.csv"'
    
    writer = csv.writer(response)
    projects = InterFep.objects.select_related(
        'lunitid', 'collabagenid', 'partship', 'projno', 'cmmsrvyid', 
        'ifpartofprogid', 'projid', 'awrdrcvdid'
    ).all()
    
    headers = [
        'Internal Project ID', 'Partnership ID', 'Project Number', 'Code',
        'Lead Unit Department', 'Lead Unit Contact Person', 'Lead Unit Contact Info',
        'Curriculum Offering', 'Collaborating Agency Name',
        'Date Application/Board Resolution',
        'Community Survey Date Conducted', 'Community Survey Stakeholder Meeting', 'Community Survey Date Presented', 'Community Survey Date Received',
        'Program Title', 'Program Lead', 'Program Approved Budget', 'Program Counterpart Fund', 'Program Total Budget',
        'Project Title', 'Project Team Member Roles', 'Project Approved Budget', 'Project Counterpart Fund', 'Project Total Budget',
        'Award Title', 'Award Conferring Agency', 'Award Date',
        'SDG Number', 'Thematic Area', 'Remarks'
    ]
    writer.writerow(headers)
    
    for proj in projects:
        writer.writerow([
            proj.interfepid, proj.partship.partship_id if proj.partship else '',
            proj.projno.project_no if proj.projno else '', proj.code or '',
            proj.lunitid.department if proj.lunitid else '', proj.lunitid.contper if proj.lunitid else '', proj.lunitid.continf if proj.lunitid else '',
            proj.curricoffid or '', proj.collabagenid.nameagen if proj.collabagenid else '',
            proj.dateappborid or '',
            proj.cmmsrvyid.datecondt if proj.cmmsrvyid else '', proj.cmmsrvyid.dvstake if proj.cmmsrvyid else '', proj.cmmsrvyid.datepresent if proj.cmmsrvyid else '', proj.cmmsrvyid.datercvd if proj.cmmsrvyid else '',
            proj.ifpartofprogid.title if proj.ifpartofprogid else '', proj.ifpartofprogid.proglead if proj.ifpartofprogid else '', proj.ifpartofprogid.apprvbudg if proj.ifpartofprogid else '', proj.ifpartofprogid.cntptfund if proj.ifpartofprogid else '', proj.ifpartofprogid.totalbudg if proj.ifpartofprogid else '',
            proj.projid.title if proj.projid else '', proj.projid.tmbrroles if proj.projid else '', proj.projid.apprvbudg if proj.projid else '', proj.projid.cntptfund if proj.projid else '', proj.projid.totalbudg if proj.projid else '',
            proj.awrdrcvdid.titleaward if proj.awrdrcvdid else '', proj.awrdrcvdid.conferagn if proj.awrdrcvdid else '', proj.awrdrcvdid.date if proj.awrdrcvdid else '',
            proj.sdgnum or '', proj.themarea or '', proj.remarks or ''
        ])
    
    return response

def export_external_projects(request):
    """Export external projects to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="external_projects.csv"'
    
    writer = csv.writer(response)
    projects = ExterFep.objects.select_related(
        'lunitid', 'collabgenid', 'partship', 'projno', 'ifpoprogid', 
        'projid', 'awardrcvdid'
    ).all()
    
    headers = [
        'External Project ID', 'Partnership ID', 'Project Number', 'Code',
        'Lead Unit Department', 'Lead Unit Contact Person', 'Lead Unit Contact Info',
        'Curriculum Offering', 'Collaborating Agency Name',
        'Program Title', 'Program Leader', 'Program Approved Budget from Funding Agency', 'Program Counterpart Fund from CVSU', 'Program Total Budget', 'Program Inclusive Dates',
        'Project Title', 'Project Team Members', 'Project Approved Budget from Funding Agency', 'Project Counterpart Fund from CVSU', 'Project Total Budget', 'Project Inclusive Dates',
        'Funding Agency', 'Date Approved by Funding Agency', 'Date of Inception Meeting',
        'SDG Number', 'Thematic Area',
        'Award Title', 'Award Conferring Agency', 'Award Date',
        'Remarks'
    ]
    writer.writerow(headers)
    
    for proj in projects:
        writer.writerow([
            proj.exterfep_id, proj.partship.partship_id if proj.partship else '',
            proj.projno.project_no if proj.projno else '', proj.code or '',
            proj.lunitid.department if proj.lunitid else '', proj.lunitid.contper if proj.lunitid else '', proj.lunitid.continf if proj.lunitid else '',
            proj.curricoffid or '', proj.collabgenid.nameagen if proj.collabgenid else '',
            proj.ifpoprogid.title if proj.ifpoprogid else '', proj.ifpoprogid.proglead if proj.ifpoprogid else '', proj.ifpoprogid.abfundagency if proj.ifpoprogid else '', proj.ifpoprogid.cpfundcvsu if proj.ifpoprogid else '', proj.ifpoprogid.tbudget if proj.ifpoprogid else '', proj.ifpoprogid.incldates if proj.ifpoprogid else '',
            proj.projid.title if proj.projid else '', proj.projid.tmembers if proj.projid else '', proj.projid.abfundagency if proj.projid else '', proj.projid.cpfundcvsu if proj.projid else '', proj.projid.tbudget if proj.projid else '', proj.projid.incldates if proj.projid else '',
            proj.fundagency or '', proj.dtappfagency or '', proj.dtinmeet or '',
            proj.sdgnum or '', proj.themarea or '',
            proj.awardrcvdid.toaward if proj.awardrcvdid else '', proj.awardrcvdid.conferagency if proj.awardrcvdid else '', proj.awardrcvdid.date if proj.awardrcvdid else '',
            proj.remarks or ''
        ])
    
    return response

def export_advisory_services(request):
    """Export advisory services to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="advisory_services.csv"'
    
    writer = csv.writer(response)
    services = AdvServices.objects.select_related(
        'lunitid', 'collabagenciesid', 'partship', 'clientsid', 
        'adservoverall', 'adservtimeliness', 'estexsourcefund'
    ).all()
    
    headers = [
        'Advisory Service ID', 'Partnership ID', 'Advisory Services Number', 'Code',
        'Lead Unit Department', 'Lead Unit Contact Person', 'Lead Unit Contact Info',
        'Curriculum Offering', 'Collaborating Agency Name',
        'Client Name', 'Client Organization', 'Client Sex', 'Client Category',
        'Advisory Services Provided', 'Inclusive Dates', 'Venue',
        'Overall Rating 1', 'Overall Rating 2', 'Overall Rating 3', 'Overall Rating 4', 'Overall Rating 5',
        'Timeliness Rating 1', 'Timeliness Rating 2', 'Timeliness Rating 3', 'Timeliness Rating 4', 'Timeliness Rating 5',
        'Fund Amount Charged CVSU', 'Fund Amount Charged Partner Agency', 'Fund Partner Agency Name',
        'Total Advisory Service Requests', 'Total Advisory Service Requests Responded',
        'SDG Number', 'Thematic Area', 'Remarks'
    ]
    writer.writerow(headers)
    
    for service in services:
        writer.writerow([
            service.advservices_id, service.partship.partship_id if service.partship else '',
            service.number or '', service.code or '',
            service.lunitid.department if service.lunitid else '', service.lunitid.contper if service.lunitid else '', service.lunitid.continf if service.lunitid else '',
            service.curricoff or '', service.collabagenciesid.nameagen if service.collabagenciesid else '',
            service.clientsid.name if service.clientsid else '', service.clientsid.organization if service.clientsid else '', service.clientsid.sex if service.clientsid else '', service.clientsid.categ if service.clientsid else '',
            service.adservprov or '', service.incldates or '', service.venue or '',
            service.adservoverall.one if service.adservoverall else '', service.adservoverall.two if service.adservoverall else '', service.adservoverall.three if service.adservoverall else '', service.adservoverall.four if service.adservoverall else '', service.adservoverall.five if service.adservoverall else '',
            service.adservtimeliness.one if service.adservtimeliness else '', service.adservtimeliness.two if service.adservtimeliness else '', service.adservtimeliness.three if service.adservtimeliness else '', service.adservtimeliness.four if service.adservtimeliness else '', service.adservtimeliness.five if service.adservtimeliness else '',
            service.estexsourcefund.amntchcvsu if service.estexsourcefund else '', service.estexsourcefund.amntchpartagency if service.estexsourcefund else '', service.estexsourcefund.partagencyname if service.estexsourcefund else '',
            service.totaladservreq or '', service.totaladservreqrespond or '',
            service.sdgnum or '', service.themarea or '', service.remarks or ''
        ])
    
    return response

def export_all_data(request):
    """Export all form data to Excel with filled and blank forms"""
    return export_forms_to_excel(request)





def advisory_services_list(request):
    """Advisory services list view"""
    services = AdvServices.objects.select_related('partship').all()
    paginator = Paginator(services, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'title': 'Advisory Service Records'
    }
    return render(request, 'advisory_services_list.html', context)

def select_form_type(request, partnership_id):
    """View to select which form to create based on partnership nature"""
    try:
        partnership = Partnership.objects.get(pk=partnership_id)
        
        available_forms = []
        if partnership.natpart in ['Internal', 'Both']:
            available_forms.append({
                'name': 'Internal Project',
                'url': reverse('internal_project_form') + f'?partnership_id={partnership_id}',
                'description': 'Create an internally funded project'
            })
        
        if partnership.natpart in ['External', 'Both']:
            available_forms.append({
                'name': 'External Project', 
                'url': reverse('external_project_form') + f'?partnership_id={partnership_id}',
                'description': 'Create an externally funded project'
            })
        
        if partnership.natpart == 'Advisory':
            available_forms.append({
                'name': 'Advisory Service',
                'url': reverse('advisory_services_form') + f'?partnership_id={partnership_id}',
                'description': 'Create an advisory service'
            })
        
        context = {
            'partnership': partnership,
            'available_forms': available_forms,
            'title': 'Select Form Type'
        }
        return render(request, 'select_form_type.html', context)
        
    except Partnership.DoesNotExist:
        messages.error(request, 'Partnership not found')
        return redirect('partnership_list')