# media_features/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db import transaction 
from django.contrib.auth.decorators import login_required  # <-- Import login_required

from openpyxl import Workbook 
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from .forms import ExtensionPPAFeaturedForm, TechnologyForm, StudentExtensionInvolvementForm, FacultyInvolvementForm
from .models import ExtensionPPAFeatured, ExtensionPPA, MediaOutlet, SupportingDocument, StudentExtensionInvolvement
from .models import Technology, Department, CurricularOffering, FormSubmission, FacultyInvolvement
from adopters_features.models import (
    Table5Adopter, Table6IEC, Table7aBudgetGAA, Table7bBudgetIncome
)
import json # Import json to handle the form_data JSONField
# from your_app_name.models import TechnologyCommercialized # Example for Table 11

def reports_dashboard(request):
    """
    Renders the main dashboard page with links to all forms.
    """
    forms = [
        {'name': 'Table 8: Faculty Involvement in ESCE', 'url_name': 'table_8_form'},
        {'name': 'Table 9: Student Involvement in ESCE', 'url_name': 'table_9_form'},
        {'name': 'Table 10: ES Activities Featured Forms', 'url_name': 'table_10_form'},
        {'name': 'Table 11: Technologies Commercialized', 'url_name': 'table_11_form'},
        
    ]
    return render(request, 'media_features/dashboard.html', {'forms': forms})

@login_required
def media_feature_form(request):
    """
    Handles the display and submission of the media feature form.
    """
    if request.method == 'POST':
        form = ExtensionPPAFeaturedForm(request.POST, request.FILES)
        if form.is_valid():
            
            with transaction.atomic():
                existing_ppa = form.cleaned_data.get('extension_ppa')
                new_ppa_name = form.cleaned_data.get('ppa_name')
                
                if new_ppa_name:
                    ppa_instance, created = ExtensionPPA.objects.get_or_create(ppa_name=new_ppa_name)
                else:
                    ppa_instance = existing_ppa

                media_outlet_name = form.cleaned_data.get('media_outlet_name')
                media_outlet_instance = None
                if media_outlet_name:
                    media_outlet_instance, created = MediaOutlet.objects.get_or_create(media_outlet_name=media_outlet_name)
                
                feature_instance = form.save(commit=False)
                feature_instance.extension_ppa = ppa_instance
                feature_instance.media_outlet = media_outlet_instance
                feature_instance.save()
                
                # List to store data for the FormSubmission JSON field
                document_data = []

                files = request.FILES.getlist('files')
                for file in files:
                    doc = SupportingDocument.objects.create(
                        extension_ppa_featured=feature_instance, 
                        file=file
                    )
                    document_data.append({
                        'name': doc.file.name.split('/')[-1],
                        'url': doc.file.url
                    })
                
                form_data_dict = {
                    'extension_ppa': ppa_instance.ppa_name if ppa_instance else None,
                    'media_outlet_name': media_outlet_instance.media_outlet_name if media_outlet_instance else None,
                    'date_featured': str(feature_instance.date_featured),
                    'remarks': feature_instance.remarks,
                    'supporting_documents': document_data # <-- ADDED
                }
                
                FormSubmission.objects.create(
                    submitter=request.user,
                    form_name='Table 10: ES Activities Features',
                    form_data=form_data_dict,
                    form_instance_id=feature_instance.pk
                )
            
            messages.success(request, 'Media feature record and documents added successfully!')
            return redirect('table_10_form')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = ExtensionPPAFeaturedForm()
    
    media_features = ExtensionPPAFeatured.objects.select_related('extension_ppa', 'media_outlet').prefetch_related('supporting_documents').all().order_by('-date_featured')

    context = {
        'form': form,
        'media_features': media_features
    }
    return render(request, 'media_features/table_10_form.html', context)

@login_required
def technology_commercialized_form(request):
    """
    Handles the display and submission of the technologies commercialized form.
    """
    if request.method == 'POST':
        form = TechnologyForm(request.POST)
        if form.is_valid():
            technology_instance = form.save()
            
            # List to store data for the FormSubmission JSON field
            document_data = []

            # Handle multiple file uploads
            files = request.FILES.getlist('files')
            for file in files:
                doc = SupportingDocument.objects.create(
                    technology=technology_instance, 
                    file=file
                )
                document_data.append({
                    'name': doc.file.name.split('/')[-1],
                    'url': doc.file.url
                })
            
            form_data_dict = {
                'department': technology_instance.department.department_name,
                'technology_title': technology_instance.technology_title,
                'year_developed': technology_instance.year_developed,
                'technology_generator': technology_instance.technology_generator,
                'technology_status': technology_instance.technology_status.status_name,
                'remarks': technology_instance.remarks,
                'curricular_offerings': [offering.offering_name for offering in technology_instance.curricular_offerings.all()],
                'supporting_documents': document_data # <-- ADDED
            }

            FormSubmission.objects.create(
                submitter=request.user,
                form_name='Table 11: Technologies Commercialized',
                form_data=form_data_dict,
                form_instance_id=technology_instance.pk
            )
            
            messages.success(request, 'Technology record and documents added successfully!')
            return redirect('table_11_form')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = TechnologyForm()
    
    technologies = Technology.objects.select_related('department').prefetch_related('curricular_offerings', 'supporting_documents').all().order_by('-year_developed')
    
    context = {
        'form': form,
        'technologies': technologies
    }
    return render(request, 'media_features/table_11_form.html', context)

def get_curricular_offerings(request, department_id):
    """
    Returns a JSON response of curricular offerings for a given department.
    """
    offerings = CurricularOffering.objects.filter(department_id=department_id).values('curricular_offering_id', 'offering_name')
    return JsonResponse(list(offerings), safe=False)


# NEW: This is the main, centralized export function
def export_all_to_excel(request):
    """
    Exports data from all specified models to a single Excel file,
    with each model on a separate worksheet.
    """
    workbook = Workbook()
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=all_extension_reports.xlsx'

    # Get the base URL for supporting document links
    base_url = request.build_absolute_uri('/')[:-1]

    # Dynamically add methods to models for easy data retrieval
    def get_supporting_documents(self):
        # A utility function to retrieve supporting documents based on the related_name
        # The related name is derived from the model name in lowercase.
        model_name = self.__class__.__name__.lower()
        if model_name == 'facultyinvolvement':
            related_name = 'faculty_involvement'
        elif model_name == 'studentextensioninvolvement':
            related_name = 'student_involvement'
        elif model_name == 'extensionppafeatured':
            related_name = 'extension_ppa_featured'
        elif model_name == 'technology':
            related_name = 'technology'
        
        docs = SupportingDocument.objects.filter(**{related_name: self})
        return ", ".join([f"{base_url}{doc.file.url}" for doc in docs])

    

    def get_curricular_offerings(self):
        return ", ".join([offering.offering_name for offering in self.curricular_offerings.all()])

    FacultyInvolvement.add_to_class('supporting_documents_list', get_supporting_documents)
    ExtensionPPAFeatured.add_to_class('supporting_documents_list', get_supporting_documents)
    Technology.add_to_class('supporting_documents_list', get_supporting_documents)
    Technology.add_to_class('curricular_offerings_list', get_curricular_offerings)
    StudentExtensionInvolvement.add_to_class('supporting_documents_list', get_supporting_documents)

    # Dictionary defining the tables to export and their configurations
    tables = {
        'Table 5 - Adopters': {
        'model': Table5Adopter,
        'fields': [
            'no',
            'code',
            'lead_unit',
            'related_curricular_offering',
            'adopter_name',
            'adopter_address',
            'adopter_contact',
            'adopter_sex',
            'adopter_category',
            'projects_involved',
            'trainings_attended',
            'other_assistance_received',
            'date_started',
            'technologies_adopted',
            'monthly_income_before',
            'monthly_income_after',
            'income_difference',
            'other_significant_changes',
            'remarks',
            'department_unit',
            'contact_person',
            'contact_number_email',
        ],
        'headers': [
            'No.',
            'Code',
            'Lead Unit',
            'Related Curricular Offering',
            'Adopter Name',
            'Adopter Address',
            'Adopter Contact',
            'Adopter Sex',
            'Adopter Category',
            'Projects Involved',
            'Trainings Attended',
            'Other Assistance Received',
            'Date Started',
            'Technologies Adopted',
            'Monthly Income Before',
            'Monthly Income After',
            'Income Difference',
            'Other Significant Changes',
            'Remarks',
            'Department/Unit',
            'Contact Person',
            'Contact Number/Email',
        ],
        'notes': ['',] * 22, # Adjust the number of notes to match the number of fields
    },
    'Table 6 - IEC': {
        'model': Table6IEC,
        'fields': [
            'no',
            'code',
            'lead_unit',
            'related_curricular_offering',
            'title',
            'format',
            'male_recipients',
            'female_recipients',
            'student_recipients',
            'farmer_recipients',
            'fisherfolk_recipients',
            'ag_technician_recipients',
            'gov_employee_recipients',
            'private_employee_recipients',
            'others_recipients',
            'total_recipients',
            'project_no',
            'sdg',
            'thematic_area',
            'remarks',
            'department_unit',
            'contact_person',
            'contact_number_email',
        ],
        'headers': [
            'No.',
            'Code',
            'Lead Unit',
            'Related Curricular Offering',
            'Title of IEC Material',
            'Format',
            'Male Recipients',
            'Female Recipients',
            'Student Recipients',
            'Farmer Recipients',
            'Fisherfolk Recipients',
            'Agricultural Technician Recipients',
            'Government Employee Recipients',
            'Private Employee Recipients',
            'Others',
            'Total Recipients',
            'Project No.',
            'SDG',
            'Thematic Area',
            'Remarks',
            'Department/Unit',
            'Contact Person',
            'Contact No./Email',
        ],
        'notes': [''] * 23,
    },
     'Table 7a - Budget (GAA)': {
        'model': Table7aBudgetGAA,
        'fields': [
            'total_budget_allocated',
            'department',
            'curricular_offering',
            'allocated_budget',
            'amount_utilized',
            'remarks',
        ],
        'headers': [
            'Total Budget Allocated for the Campus/College/Unit',
            'Department',
            'Curricular Offering',
            'Allocated Budget for Extension per Department',
            'Amount Utilized for the period',
            'Remarks',
        ],
        'notes': [
            'Php',
            '',
            '',
            'If there\'s any',
            'Disbursed and Accounts Payable',
            '',
        ],
    },
    'Table 7b - Budget (Income)': {
        'model': Table7bBudgetIncome,
        'fields': [
            'total_budget_allocated',
            'department',
            'curricular_offering',
            'allocated_budget',
            'amount_utilized',
            'remarks',
        ],
        'headers': [
            'Total Budget Allocated for the Campus/College/Unit',
            'Department',
            'Curricular Offering',
            'Allocated Budget for Extension per Department',
            'Amount Utilized for the period',
            'Remarks',
        ],
        'notes': [
            'Php',
            '',
            '',
            'If there\'s any',
            'Disbursed and Accounts Payable',
            '',
        ],
    },
        'Table 8 - Faculty Involvement': {
            'model': FacultyInvolvement,
            'fields': [
                'faculty_staff_name',
                'academic_rank_position', 
                'employment_status',
                'avg_hours_per_week',
                'total_hours_per_quarter',
                'remarks',
                'supporting_documents_list'
            ],
            'headers': [
                'Faculty/Staff Name',
                'Academic Rank/Position',
                'Employment Status', 
                'Average Hours per Week',
                'Total Hours per Quarter',
                'Remarks',
                'Supporting Documents'
            ],
            'notes': ['', '', '', '', '', '', '*Supporting documentation for faculty involvement']
        },
        'Table 9 - Student Involvement': {
            'model': StudentExtensionInvolvement,
            'fields': [
                'department__department_name',
                'curricular_offering__offering_name',
                'total_students_for_period',
                'students_involved_in_extension',
                'percentage_students_involved',
                'remarks',
                'supporting_documents_list'
            ],
            'headers': [
                'Department',
                'Curricular Offering',
                'Total Number of Students for the period (a)',
                'Number of Students involved in extension activities (b)',
                'Percentage of Students involved (%)',
                'Remarks',
                'Supporting Documents'
            ],
            'notes': ['', '', '', '', '', '', '1. List of students involved per extension activity\n2. Photo documentation']
        },
        'Table 10 - Media Features': {
            'model': ExtensionPPAFeatured,
            'fields': [
                'extension_ppa__ppa_name',
                'media_outlet__media_outlet_name',
                'date_featured',
                'remarks',
                'supporting_documents_list'
            ],
            'headers': [
                'Extension Program/Project/Activity (PPA)',
                'Print, radio, online media where Extension PPA was featured',
                'Date Featured',
                'Remarks',
                'Supporting Documents'
            ],
            'notes': ['', '', '', '', '*Copy of any evidence showing the Extension PPA was featured']
        },
        'Table 11 - Technologies': {
            'model': Technology,
            'fields': [
                'department__department_name',
                'technology_title',
                'year_developed',
                'technology_generator',
                'technology_status__status_name',
                'curricular_offerings_list',
                'remarks',
                'supporting_documents_list'
            ],
            'headers': [
                'Department',
                'Technology Title',
                'Year Developed',
                'Technology Generator',
                'Technology Status',
                'Related Curricular Offerings',
                'Remarks',
                'Supporting Documents'
            ],
            'notes': ['', '', '', '', '', '', '', '1. Photo and IEC material about the technology\n2. Documentation of deployment, commercialization, and pre-commercialization activities']
        },
    }

    # Remove the default sheet created by openpyxl
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])

    for sheet_name, table_info in tables.items():
        model = table_info['model']
        fields = table_info['fields']
        headers = table_info['headers']
        notes = table_info['notes']

        sheet = workbook.create_sheet(title=sheet_name)
        
        # Write headers and notes
        for col_num, header_text in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header_text)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical='center')

        for col_num, note_text in enumerate(notes, 1):
            if note_text:
                cell = sheet.cell(row=2, column=col_num, value=note_text)
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        sheet.row_dimensions[1].height = 40
        sheet.row_dimensions[2].height = 60
        
        # Build the queryset with optimized lookups
        queryset = model.objects.all()
        if model == ExtensionPPAFeatured:
            queryset = queryset.select_related('extension_ppa', 'media_outlet').prefetch_related('supporting_documents')
        elif model == Technology:
            queryset = queryset.select_related('department', 'technology_status').prefetch_related('curricular_offerings', 'supporting_documents')
        elif model == StudentExtensionInvolvement:
            queryset = queryset.select_related('department', 'curricular_offering').prefetch_related('supporting_documents')

        # Write data rows
        start_row = 3
        for record in queryset:
            row_data = []
            for field in fields:
                try:
                    # Use the dynamically added methods for M2M and FileFields
                    if field in ['curricular_offerings_list', 'supporting_documents_list']:
                        value = getattr(record, field)()
                    else:
                        # Use split to handle lookups like department__department_name
                        parts = field.split('__')
                        value = record
                        for part in parts:
                            value = getattr(value, part)
                except (AttributeError, TypeError):
                    value = ''  # Handle cases where a related object might be null
                
                row_data.append(value)
            
            for col_num, cell_value in enumerate(row_data, 1):
                cell = sheet.cell(row=start_row, column=col_num, value=cell_value)
                cell.alignment = Alignment(wrap_text=True)
            start_row += 1

    workbook.save(response)
    return response

@login_required
def student_involvement_form(request):
    if request.method == 'POST':
        form = StudentExtensionInvolvementForm(request.POST)
        files = request.FILES.getlist('files') # Get the uploaded files
        if form.is_valid():
            with transaction.atomic():
                # Save the main form instance first
                involvement_instance = form.save()

                # List to store data for the FormSubmission JSON field
                document_data = []
                
                # Handle file uploads and create SupportingDocument records
                for f in files:
                    doc = SupportingDocument.objects.create(
                        student_involvement=involvement_instance,
                        file=f,
                        submitter=request.user
                    )
                    # Add the document's URL and name to our list
                    document_data.append({
                        'name': doc.file.name.split('/')[-1], # Use split to get just the filename
                        'url': doc.file.url
                    })

                # Create the FormSubmission record
                FormSubmission.objects.create(
                    submitter=request.user,
                    form_name='Table 9: Student Involvement in ESCE',
                    # Store a copy of the form data AND the document data
                    form_data={
                        'department': str(involvement_instance.department),
                        'curricular_offering': str(involvement_instance.curricular_offering),
                        'total_students_for_period': involvement_instance.total_students_for_period,
                        'students_involved_in_extension': involvement_instance.students_involved_in_extension,
                        'percentage_students_involved': involvement_instance.percentage_students_involved,
                        'remarks': involvement_instance.remarks,
                        'supporting_documents': document_data # <-- ADDED: list of document info
                    },
                    form_instance_id=involvement_instance.pk
                )
            
            messages.success(request, 'Form submitted successfully!')
            return redirect('table_9_form')
    else:
        form = StudentExtensionInvolvementForm()
    
    student_involvements = StudentExtensionInvolvement.objects.all()
    context = {
        'form': form,
        'student_involvements': student_involvements
    }
    return render(request, 'media_features/table_9_form.html', context)



def faculty_involvement_form(request):
    """
    Handles the display and submission of the faculty involvement form (Table 8).
    """
    if request.method == 'POST':
        form = FacultyInvolvementForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                with transaction.atomic():
                    instance = form.save(commit=False)
                    instance.submitter = request.user
                    instance.save()

                    document_data = []
                    supporting_documents = request.FILES.getlist('files') # Changed to 'files' to match your HTML
                    for doc_file in supporting_documents:
                        doc = SupportingDocument.objects.create(
                            faculty_involvement=instance,
                            file=doc_file,
                            submitter=request.user,
                        )
                        document_data.append({
                            'name': doc.file.name.split('/')[-1],
                            'url': doc.file.url
                        })

                    # Create FormSubmission record with the data and document list
                    form_data_dict = {
                        'faculty_staff_name': instance.faculty_staff_name,
                        'academic_rank_position': instance.academic_rank_position,
                        'employment_status': instance.employment_status,
                        'avg_hours_per_week': float(instance.avg_hours_per_week),
                        'total_hours_per_quarter': float(instance.total_hours_per_quarter),
                        'remarks': instance.remarks,
                        'supporting_documents': document_data # Now correctly includes the document list
                    }
                    FormSubmission.objects.create(
                        submitter=request.user,
                        form_name='Table 8: Faculty Involvement in ESCE',
                        form_data=form_data_dict,
                        form_instance_id=instance.id
                    )
               
                messages.success(request, 'Faculty involvement record and documents added successfully!')
                return redirect('table_8_form')

            except Exception as e:
                messages.error(request, f'An error occurred: {e}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = FacultyInvolvementForm()

    records = FacultyInvolvement.objects.prefetch_related('supporting_documents').all().order_by('-submitted_at')
    context = {
        'form': form,
        'records': records,
    }
    return render(request, 'media_features/table_8_form.html', context)


@login_required
def faculty_involvement_details(request, submission_id):
    submission = get_object_or_404(FormSubmission, pk=submission_id)
    # This view no longer needs to use form_instance_id for documents
    context = { 'submission': submission, 'form_data': submission.form_data }
    return render(request, 'media_features/table_8_details.html', context)

@login_required
def student_involvement_details(request, submission_id):
    submission = get_object_or_404(FormSubmission, pk=submission_id)
    form_data = get_object_or_404(StudentExtensionInvolvement, pk=submission.form_id)
    context = { 'submission': submission, 'form_data': form_data }
    return render(request, 'media_features/table_9_details.html', context)

@login_required
def media_features_details(request, submission_id):
    submission = get_object_or_404(FormSubmission, pk=submission_id)
    form_data = get_object_or_404(ExtensionPPAFeatured, pk=submission.form_id)
    context = { 'submission': submission, 'form_data': form_data }
    return render(request, 'media_features/table_10_details.html', context)

@login_required
def technologies_details(request, submission_id):
    submission = get_object_or_404(FormSubmission, pk=submission_id)
    form_data = get_object_or_404(Technology, pk=submission.form_id)
    context = { 'submission': submission, 'form_data': form_data }
    return render(request, 'media_features/table_11_details.html', context)