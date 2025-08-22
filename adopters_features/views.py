from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from .forms import Table5Form, Table6Form, Table7aForm, Table7bForm
from .models import Table5Adopter, Table6IEC, Table7aBudgetGAA, Table7bBudgetIncome, Submission
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os

# ------------- Dashboard -------------
def dashboard(request):
    """
    Renders the dashboard page.
    """
    return render(request, 'adopters_features/dashboard.html')

# ---------- Helpers (Excel) ----------
def queryset_to_sheet(ws, queryset, field_order=None, header_map=None):
    """
    Populates an openpyxl worksheet with data from a Django queryset.
    
    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): The worksheet to populate.
        queryset (django.db.models.query.QuerySet): The data to write.
        field_order (list, optional): A list of field names to determine column order.
                                     Defaults to all fields on the model.
        header_map (dict, optional): A dictionary to map field names to custom headers.
                                     Defaults to converting field names to titles.
    """
    if field_order is None:
        model = queryset.model
        field_order = [f.name for f in model._meta.fields]

    headers = []
    for f in field_order:
        pretty = header_map.get(f, f.replace('_', ' ').title()) if header_map else f.replace('_', ' ').title()
        headers.append(pretty)
    ws.append(headers)

    for obj in queryset:
        row = [getattr(obj, f) for f in field_order]
        ws.append(row)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].auto_size = True

def wb_response(wb, filename):
    """
    Creates an HTTP response for downloading an Excel workbook.
    """
    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp

def handle_delete(request, pk, model_class, redirect_name):
    """
    Generic helper function to handle object deletion.
    """
    obj = get_object_or_404(model_class, pk=pk)
    if request.method == 'POST':
        obj.delete()
        return redirect(redirect_name)

# ---------- Table 5 ----------
def table5_view(request):
    """
    Handles form submission and displays existing data for Table 5.
    
    The 'no' field is now handled in the template to avoid race conditions.
    """
    if request.method == 'POST':
        form = Table5Form(request.POST, request.FILES)
        if form.is_valid():
            # Create a submission instance and link the form data to it
            submission = Submission.objects.create(submitter=request.user)
            table5_instance = form.save(commit=False)
            table5_instance.submission = submission
            
            # The 'no' field is no longer set here. It's safer to handle numbering
            # on the front end or rely on the unique database ID.
            table5_instance.save()
            return redirect('table5')
    else:
        form = Table5Form()
    
    # Retrieve all rows and order by ID for consistent numbering
    rows = Table5Adopter.objects.all().order_by('id')
    
    # Pass the form and rows to the template
    return render(request, 'adopters_features/Table_5_form.html', {'form': form, 'rows': rows})

def table5_edit(request, pk):
    """
    Handles editing an existing entry for Table 5.
    """
    obj = get_object_or_404(Table5Adopter, pk=pk)
    if request.method == 'POST':
        form = Table5Form(request.POST, request.FILES, instance=obj)
        if form.is_valid():
            form.save()
            return redirect('table5')
    else:
        form = Table5Form(instance=obj)
    return render(request, 'adopters_features/Table_5_edit_form.html', {'form': form})

def table5_delete(request, pk):
    """
    Handles deleting an entry for Table 5 using the helper function.
    """
    return handle_delete(request, pk, Table5Adopter, 'table5')

def export_table5_excel(request):
    """
    Exports all Table 5 data to an Excel file.
    """
    qs = Table5Adopter.objects.all().order_by('id')
    wb = Workbook(); ws = wb.active; ws.title = "Table 5 - Adopters"
    queryset_to_sheet(ws, qs)
    return wb_response(wb, 'table5_adopters.xlsx')

def table5_details(request, submission_id):
    """
    Views details of a specific Table 5 submission.
    """
    submission = get_object_or_404(Submission, id=submission_id)
    form_data = get_object_or_404(Table5Adopter, submission=submission)
    context = {
        'submission': submission,
        'form_data': form_data,
    }
    return render(request, 'adopters_features/Table_5_details.html', context)

# ---------- Table 6 ----------
def table6_view(request):
    """
    Handles form submission and displays existing data for Table 6.
    """
    if request.method == 'POST':
        form = Table6Form(request.POST, request.FILES)
        if form.is_valid():
            submission = Submission.objects.create(submitter=request.user)
            table6_instance = form.save(commit=False)
            table6_instance.submission = submission
            table6_instance.save()
            return redirect('table6')
    else:
        form = Table6Form()
    
    rows = Table6IEC.objects.all().order_by('id')
    return render(request, 'adopters_features/Table_6_form.html', {'form': form, 'rows': rows})

def table6_edit(request, pk):
    """
    Handles editing an existing entry for Table 6.
    """
    obj = get_object_or_404(Table6IEC, pk=pk)
    if request.method == 'POST':
        form = Table6Form(request.POST, request.FILES, instance=obj)
        if form.is_valid():
            form.save()
            return redirect('table6')
    else:
        form = Table6Form(instance=obj)
    return render(request, 'adopters_features/Table_6_edit_form.html', {'form': form})

def table6_delete(request, pk):
    """
    Handles deleting an entry for Table 6.
    """
    return handle_delete(request, pk, Table6IEC, 'table6')

def export_table6_excel(request):
    """
    Exports all Table 6 data to an Excel file.
    """
    qs = Table6IEC.objects.all().order_by('id')
    wb = Workbook(); ws = wb.active; ws.title = "Table 6 - IEC"
    queryset_to_sheet(ws, qs)
    return wb_response(wb, 'table6_iec.xlsx')

def table6_details(request, submission_id):
    """
    Views details of a specific Table 6 submission.
    """
    submission = get_object_or_404(Submission, id=submission_id)
    form_data = get_object_or_404(Table6IEC, submission=submission)
    context = {
        'submission': submission,
        'form_data': form_data,
    }
    return render(request, 'adopters_features/Table_6_details.html', context)

# ---------- Table 7a ----------
def table7a_view(request):
    """
    Handles form submission and displays existing data for Table 7a.
    """
    if request.method == 'POST':
        form = Table7aForm(request.POST, request.FILES)
        if form.is_valid():
            submission = Submission.objects.create(submitter=request.user)
            table7a_instance = form.save(commit=False)
            table7a_instance.submission = submission
            table7a_instance.save()
            return redirect('table7a')
    else:
        form = Table7aForm()
    
    rows = Table7aBudgetGAA.objects.all().order_by('id')
    return render(request, 'adopters_features/Table_7a_form.html', {'form': form, 'rows': rows})

def table7a_edit(request, pk):
    """
    Handles editing an existing entry for Table 7a.
    """
    obj = get_object_or_404(Table7aBudgetGAA, pk=pk)
    if request.method == 'POST':
        form = Table7aForm(request.POST, request.FILES, instance=obj)
        if form.is_valid():
            form.save()
            return redirect('table7a')
    else:
        form = Table7aForm(instance=obj)
    return render(request, 'adopters_features/Table_7a_edit_form.html', {'form': form})

def table7a_delete(request, pk):
    """
    Handles deleting an entry for Table 7a.
    """
    return handle_delete(request, pk, Table7aBudgetGAA, 'table7a')

def export_table7a_excel(request):
    """
    Exports all Table 7a data to an Excel file.
    """
    qs = Table7aBudgetGAA.objects.all().order_by('id')
    wb = Workbook(); ws = wb.active; ws.title = "Table 7a - GAA"
    queryset_to_sheet(ws, qs)
    return wb_response(wb, 'table7a_budget_gaa.xlsx')

def table7a_details(request, submission_id):
    """
    Views details of a specific Table 7a submission.
    """
    submission = get_object_or_404(Submission, id=submission_id)
    form_data = get_object_or_404(Table7aBudgetGAA, submission=submission)
    context = {
        'submission': submission,
        'form_data': form_data,
    }
    return render(request, 'adopters_features/Table_7a_details.html', context)

# ---------- Table 7b ----------
def table7b_view(request):
    """
    Handles form submission and displays existing data for Table 7b.
    """
    if request.method == 'POST':
        form = Table7bForm(request.POST, request.FILES)
        if form.is_valid():
            submission = Submission.objects.create(submitter=request.user)
            table7b_instance = form.save(commit=False)
            table7b_instance.submission = submission
            table7b_instance.save()
            return redirect('table7b')
    else:
        form = Table7bForm()
    
    rows = Table7bBudgetIncome.objects.all().order_by('id')
    return render(request, 'adopters_features/Table_7b_form.html', {'form': form, 'rows': rows})

def table7b_edit(request, pk):
    """
    Handles editing an existing entry for Table 7b.
    """
    obj = get_object_or_404(Table7bBudgetIncome, pk=pk)
    if request.method == 'POST':
        form = Table7bForm(request.POST, request.FILES, instance=obj)
        if form.is_valid():
            form.save()
            return redirect('table7b')
    else:
        form = Table7bForm(instance=obj)
    return render(request, 'adopters_features/Table_7b_edit_form.html', {'form': form})

def table7b_delete(request, pk):
    """
    Handles deleting an entry for Table 7b.
    """
    return handle_delete(request, pk, Table7bBudgetIncome, 'table7b')

def export_table7b_excel(request):
    """
    Exports all Table 7b data to an Excel file.
    """
    qs = Table7bBudgetIncome.objects.all().order_by('id')
    wb = Workbook(); ws = wb.active; ws.title = "Table 7b - Income"
    queryset_to_sheet(ws, qs)
    return wb_response(wb, 'table7b_budget_income.xlsx')

def table7b_details(request, submission_id):
    """
    Views details of a specific Table 7b submission.
    """
    submission = get_object_or_404(Submission, id=submission_id)
    form_data = get_object_or_404(Table7bBudgetIncome, submission=submission)
    context = {
        'submission': submission,
        'form_data': form_data,
    }
    return render(request, 'adopters_features/Table_7b_details.html', context)

# ---------- Export ALL ----------
def export_all_excel(request):
    """
    Exports all data from all tables to a single Excel workbook with multiple sheets.
    """
    wb = Workbook()
    
    ws5 = wb.active; ws5.title = "Table 5 - Adopters"
    queryset_to_sheet(ws5, Table5Adopter.objects.all().order_by('id'))
    
    ws6 = wb.create_sheet(title="Table 6 - IEC")
    queryset_to_sheet(ws6, Table6IEC.objects.all().order_by('id'))
    
    ws7a = wb.create_sheet(title="Table 7a - GAA")
    queryset_to_sheet(ws7a, Table7aBudgetGAA.objects.all().order_by('id'))
    
    ws7b = wb.create_sheet(title="Table 7b - Income")
    queryset_to_sheet(ws7b, Table7bBudgetIncome.objects.all().order_by('id'))
    
    return wb_response(wb, 'all_tables.xlsx')

# ---------- View Submission Details (CLEANED UP) ----------
def view_submission_details(request, submission_id):
    """
    Centralized function to view submission details, using a more robust method.
    """
    submission = get_object_or_404(Submission, id=submission_id)
    
    models = [
        (Table5Adopter, 'Table_5_details.html'),
        (Table6IEC, 'Table_6_details.html'),
        (Table7aBudgetGAA, 'Table_7a_details.html'),
        (Table7bBudgetIncome, 'Table_7b_details.html'),
    ]

    for model_class, template_name in models:
        try:
            form_data = model_class.objects.get(submission=submission)
            context = {
                'submission': submission,
                'form_data': form_data,
            }
            return render(request, f'adopters_features/{template_name}', context)
        except model_class.DoesNotExist:
            continue
    
    # If we get here, no data was found for any of the models
    return render(request, 'adopters_features/no_data_found.html', {'submission': submission})
