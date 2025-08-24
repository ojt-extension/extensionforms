# accounts/views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, authenticate, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse

from .forms import OJTCoordinatorCreationForm, OJTCoordinatorLoginForm
from media_features.models import FormSubmission, ExtensionPPAFeatured, Technology, StudentExtensionInvolvement, FacultyInvolvement, SupportingDocument
from partnerships.models import Partnership, InterFep, ExterFep, AdvServices
from openpyxl.utils import get_column_letter
import openpyxl
from openpyxl.styles import Font, Alignment
from io import BytesIO

def coordinator_auth(request):
    # Redirect authenticated users to their respective dashboards
    if request.user.is_authenticated:
        if request.user.is_superuser:
            return redirect('admin_dashboard')
        return redirect('reports_dashboard')

    login_form = OJTCoordinatorLoginForm()
    signup_form = OJTCoordinatorCreationForm()
    show_signup = False

    if request.method == 'POST':
        if 'signup_submit' in request.POST:
            signup_form = OJTCoordinatorCreationForm(request.POST)
            if signup_form.is_valid():
                user = signup_form.save()
                messages.success(request, 'Account created successfully. You can now log in.')
                return redirect('coordinator_auth')
            else:
                show_signup = True  # Show the signup form again if there are errors
                messages.error(request, 'Please correct the errors in the sign-up form.')
        
        elif 'login_submit' in request.POST:
            login_form = OJTCoordinatorLoginForm(request, data=request.POST)
            if login_form.is_valid():
                user = login_form.get_user()
                login(request, user)
                if user.is_superuser:
                    return redirect('admin_dashboard')
                return redirect('reports_dashboard')
            else:
                messages.error(request, 'Invalid username or password.')

    context = {
        'login_form': login_form,
        'signup_form': signup_form,
        'show_signup': show_signup # Use this context variable to control the initial tab in JavaScript
    }
    return render(request, 'accounts/login_signup.html', context)


def coordinator_logout(request):
    logout(request)
    return redirect('coordinator_auth')

@login_required
def coordinator_dashboard(request):
    """
    The main dashboard for OJT coordinators.
    This view now passes a list of forms to the template.
    """
    forms = [
        {'name': 'Table 1: Partnerships', 'url_name': 'partnership_form'},
        {'name': 'Table 2: Internal Projects', 'url_name': 'internal_project_form'},
        {'name': 'Table 3: External Projects', 'url_name': 'external_project_form'},
        {'name': 'Table 4: Advisory Services', 'url_name': 'advisory_services_form'},
        {'name': 'Table 8: Faculty Involvement', 'url_name': 'table_8_form'},
        {'name': 'Table 9: Student Involvement', 'url_name': 'table_9_form'},
        {'name': 'Table 10: Media Features', 'url_name': 'table_10_form'},
        {'name': 'Table 11: Technologies Commercialized', 'url_name': 'table_11_form'},
    ]
    
    context = {
        'forms': forms,
    }
    return render(request, 'media_features/dashboard.html', context)

@login_required
def admin_dashboard(request):
    """
    The custom dashboard for the superuser.
    Shows a list of all submitted forms.
    """
    if not request.user.is_superuser:
        return redirect('coordinator_dashboard')
        
    submitted_forms = FormSubmission.objects.all().select_related('submitter')
    partnerships = Partnership.objects.all()
    internal_projects = InterFep.objects.all()
    external_projects = ExterFep.objects.all()
    advisory_services = AdvServices.objects.all()

    context = {
        'submitted_forms': submitted_forms,
        'partnerships': partnerships,
        'internal_projects': internal_projects,
        'external_projects': external_projects,
        'advisory_services': advisory_services,
    }
    return render(request, 'accounts/admin_dashboard.html', context)

@login_required
def export_all_submissions(request):
    """
    Exports data from all specified models to a single Excel file,
    with each model on a separate worksheet.
    """
    if not request.user.is_superuser:
        return redirect('coordinator_dashboard')

    workbook = openpyxl.Workbook()
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=all_extension_reports.xlsx'

    # Get the base URL for supporting document links
    base_url = request.build_absolute_uri('/')[:-1]

    # Dynamically add methods to models for easy data retrieval
    def get_supporting_documents(self):
        # A utility function to retrieve supporting documents based on the related_name
        model_name = self.__class__.__name__.lower()
        if model_name == 'facultyinvolvement':
            related_name = 'faculty_involvement'
        elif model_name == 'studentextensioninvolvement':
            related_name = 'student_involvement'
        elif model_name == 'extensionppafeatured':
            related_name = 'extension_ppa_featured'
        elif model_name == 'technology':
            related_name = 'technology'
        
        docs = SupportingDocument.objects.filter(**{related_name: self})
        return ", ".join([f"{base_url}{doc.file.url}" for doc in docs])

    def get_curricular_offerings(self):
        return ", ".join([offering.offering_name for offering in self.curricular_offerings.all()])

    # Add these methods to the models so the export logic can use them
    FacultyInvolvement.add_to_class('supporting_documents_list', get_supporting_documents)
    ExtensionPPAFeatured.add_to_class('supporting_documents_list', get_supporting_documents)
    Technology.add_to_class('supporting_documents_list', get_supporting_documents)
    Technology.add_to_class('curricular_offerings_list', get_curricular_offerings)
    StudentExtensionInvolvement.add_to_class('supporting_documents_list', get_supporting_documents)

    # Dictionary defining the tables to export and their configurations
    tables = {
        #added partnership, internal project, external project, and advisory service tables
        'Table 1 - Partnerships': {
            'model': Partnership,
            'fields': ['partship_id', 'moano', 'code', 'leadunit__department', 'partagency__nameagen', 'natpart', 'sdgnum', 'themarea'],
            'headers': ['Partnership ID', 'MOA Number', 'Code', 'Lead Unit', 'Partner Agency', 'Nature', 'SDG Number', 'Thematic Area'],
            'notes': ['', '', '', '', '', '', '', '']
        },
        'Table 2 - Internal Projects': {
            'model': InterFep,
            'fields': ['interfepid', 'partship__partship_id', 'code', 'lunitid__department', 'collabagenid__nameagen', 'benef', 'sdgnum'],
            'headers': ['Project ID', 'Partnership ID', 'Code', 'Lead Unit', 'Collaborating Agency', 'Beneficiaries', 'SDG Number'],
            'notes': ['', '', '', '', '', '', '']
        },
        'Table 3 - External Projects': {
            'model': ExterFep,
            'fields': ['exterfep_id', 'partship__partship_id', 'code', 'lunitid__department', 'collabgenid__nameagen', 'fundagency', 'benef'],
            'headers': ['Project ID', 'Partnership ID', 'Code', 'Lead Unit', 'Collaborating Agency', 'Funding Agency', 'Beneficiaries'],
            'notes': ['', '', '', '', '', '', '']
        },
        'Table 4 - Advisory Services': {
            'model': AdvServices,
            'fields': ['advservices_id', 'partship__partship_id', 'code', 'lunitid__department', 'adservprov', 'venue', 'totaladservreq'],
            'headers': ['Service ID', 'Partnership ID', 'Code', 'Lead Unit', 'Services Provided', 'Venue', 'Total Requests'],
            'notes': ['', '', '', '', '', '', '']
        },

        'Table 8 - Faculty Involvement': {
            'model': FacultyInvolvement,
            'fields': ['faculty_staff_name', 'academic_rank_position', 'employment_status', 'avg_hours_per_week', 'total_hours_per_quarter', 'remarks', 'supporting_documents_list'],
            'headers': ['Name of Faculty/Staff (Last, First MI.)', 'Academic Rank (Faculty)/ Position (Staff)', 'Employment Status', 'Average number of hours engaged (per week)', 'Total Number of hours engaged (per quarter)', 'Remarks', 'Supporting Documents'],
            'notes': ['', '', '', '', '', '', '1. Copy of approved schedule for the semester']
        },
        'Table 9 - Student Involvement': {
            'model': StudentExtensionInvolvement,
            'fields': ['department__department_name', 'curricular_offering__offering_name', 'total_students_for_period', 'students_involved_in_extension', 'percentage_students_involved', 'remarks', 'supporting_documents_list'],
            'headers': ['Department', 'Curricular Offering', 'Total Number of Students for the period (a)', 'Number of Students involved in extension activities (b)', 'Percentage of Students involved (%)', 'Remarks', 'Supporting Documents'],
            'notes': ['', '', '', '', '', '', '1. List of students involved per extension activity\n2. Photo documentation']
        },
        'Table 10 - Media Features': {
            'model': ExtensionPPAFeatured,
            'fields': ['extension_ppa__ppa_name', 'media_outlet__media_outlet_name', 'date_featured', 'remarks', 'supporting_documents_list'],
            'headers': ['Extension Program/Project/Activity (PPA)', 'Print, radio, online media where Extension PPA was featured', 'Date Featured', 'Remarks', 'Supporting Documents'],
            'notes': ['', '', '', '', '*Copy of any evidence showing the Extension PPA was featured']
        },
        'Table 11 - Technologies': {
            'model': Technology,
            'fields': ['department__department_name', 'technology_title', 'year_developed', 'technology_generator', 'technology_status__status_name', 'curricular_offerings_list', 'remarks', 'supporting_documents_list'],
            'headers': ['Department', 'Technology Title', 'Year Developed', 'Technology Generator', 'Technology Status', 'Related Curricular Offerings', 'Remarks', 'Supporting Documents'],
            'notes': ['', '', '', '', '', '', '', '1. Photo and IEC material about the technology\n2. Documentation of deployment, commercialization, and pre-commercialization activities']
        },
    }

    # Remove the default sheet created by openpyxl
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])

    for sheet_name, table_info in tables.items():
        model = table_info['model']
        fields = table_info['fields']
        headers = table_info['headers']
        notes = table_info['notes']

        safe_sheet_name = sheet_name.replace(':', ' -')[:31] 
        worksheet = workbook.create_sheet(title=safe_sheet_name)
        
        # Write headers and notes
        for col_num, header_text in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header_text)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical='center')

        for col_num, note_text in enumerate(notes, 1):
            if note_text:
                cell = worksheet.cell(row=2, column=col_num, value=note_text)
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        worksheet.row_dimensions[1].height = 40
        worksheet.row_dimensions[2].height = 60

        queryset = model.objects.all()
        # Optimize queryset with prefetch and select related calls
        if model == ExtensionPPAFeatured:
            queryset = queryset.select_related('extension_ppa', 'media_outlet').prefetch_related('supporting_documents')
        elif model == Technology:
            queryset = queryset.select_related('department', 'technology_status').prefetch_related('curricular_offerings', 'supporting_documents')
        elif model == StudentExtensionInvolvement:
            queryset = queryset.select_related('department', 'curricular_offering').prefetch_related('supporting_documents')
        elif model == FacultyInvolvement:
            queryset = queryset.prefetch_related('supporting_documents')
        #added partnership, internal project, external project, and advisory service query optimizations
        elif model == Partnership:
            queryset = queryset.select_related('leadunit', 'partagency')
        elif model == InterFep:
            queryset = queryset.select_related('lunitid', 'collabagenid', 'partship')
        elif model == ExterFep:
            queryset = queryset.select_related('lunitid', 'collabgenid', 'partship')
        elif model == AdvServices:
            queryset = queryset.select_related('lunitid', 'collabagenciesid', 'partship')

        
        # Write data rows
        start_row = 3
        for record in queryset:
            row_data = []
            for field in fields:
                try:
                    if field in ['curricular_offerings_list', 'supporting_documents_list']:
                        value = getattr(record, field)()
                    else:
                        parts = field.split('__')
                        value = record
                        for part in parts:
                            value = getattr(value, part)
                            if callable(value):
                                value = value()
                except (AttributeError, TypeError):
                    value = ''
                
                row_data.append(value)
            
            for col_num, cell_value in enumerate(row_data, 1):
                cell = worksheet.cell(row=start_row, column=col_num, value=cell_value)
                cell.alignment = Alignment(wrap_text=True)
            start_row += 1

    workbook.save(response)
    return response
def view_submission_details(request, submission_id):
    """
    Renders a detailed view of a single form submission.
    """
    submission = get_object_or_404(FormSubmission, pk=submission_id)

    # Determine which template to render based on the form_name
    # Add these conditions before the existing if statements:
    if   'Table 1' in submission.form_name:
        template = 'partnerships/partnership_details.html'
    elif 'Table 2' in submission.form_name:
        template = 'partnerships/internal_project_details.html'
    elif 'Table 3' in submission.form_name:
        template = 'partnerships/external_project_details.html'
    elif 'Table 4' in submission.form_name:
        template = 'partnerships/advisory_service_details.html'
    elif 'Table 10' in submission.form_name:
        template = 'media_features/table_10_details.html'
    elif 'Table 11' in submission.form_name:
        template = 'media_features/table_11_details.html'
    elif 'Table 9' in submission.form_name:
        template = 'media_features/table_9_details.html'
    elif 'Table 8' in submission.form_name:
        template = 'media_features/table_8_details.html'
    else:
        # Fallback for other forms
        template = 'media_features/generic_details.html'

    context = {
        'submission': submission,
        'form_data': submission.form_data, # Pass the JSON data directly as it contains the document URLs
    }

    return render(request, template, context)